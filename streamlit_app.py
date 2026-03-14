import streamlit as st
import pandas as pd
import datetime
from datetime import datetime, timedelta
import re
import io
import time
import json
import mimetypes
import numpy as np
import altair as alt # Thêm thư viện vẽ biểu đồ

# =============================================================================
# 1. CẤU HÌNH & KHỞI TẠO
# =============================================================================
APP_VERSION = "CRM SYSTEM"
st.set_page_config(page_title=f"CRM {APP_VERSION}", layout="wide", page_icon="💎")

# CSS UI
st.markdown("""
    <style>
    button[data-baseweb="tab"] div p { font-size: 18px !important; font-weight: 700 !important; }
    .card-3d { border-radius: 12px; padding: 20px; color: white; text-align: center; box-shadow: 0 4px 8px rgba(0,0,0,0.2); margin-bottom: 10px; }
    .bg-sales { background: linear-gradient(135deg, #00b09b, #96c93d); }
    .bg-cost { background: linear-gradient(135deg, #ff5f6d, #ffc371); }
    .bg-profit { background: linear-gradient(135deg, #f83600, #f9d423); }
    [data-testid="stDataFrame"] > div { max-height: 750px; }
    .highlight-low { background-color: #ffcccc !important; color: red !important; font-weight: bold; }
    
    /* CSS CHO CÁC NÚT BẤM: NỀN TỐI - CHỮ SÁNG */
    div.stButton > button { 
        width: 100%; 
        border-radius: 5px; 
        font-weight: bold; 
        background-color: #262730; /* Nền tối */
        color: #ffffff; /* Chữ trắng */
        border: 1px solid #4e4e4e;
    }
    div.stButton > button:hover {
        background-color: #444444;
        color: #ffffff;
        border-color: #ffffff;
    }
    
    /* STYLE CHO TOTAL VIEW */
    .total-view {
        font-size: 20px;
        font-weight: bold;
        color: #00FF00; /* Màu xanh lá nổi bật */
        background-color: #262730;
        padding: 10px;
        border-radius: 8px;
        text-align: right;
        margin-top: 10px;
        border: 1px solid #4e4e4e;
    }

    /* --- FIX: STYLE CHO DÒNG TOTAL (DÒNG CUỐI CÙNG TRONG TABLE) MÀU VÀNG --- */
    [data-testid="stDataFrame"] table tbody tr:last-child {
        background-color: #FFD700 !important; /* Màu vàng */
        color: #000000 !important; /* Chữ đen */
        font-weight: 900 !important;
    }
    [data-testid="stDataFrame"] table tbody tr:last-child td {
        color: #000000 !important;
        background-color: #FFD700 !important; /* Force nền vàng cho từng ô */
        font-weight: bold !important;
    }
    </style>""", unsafe_allow_html=True)

# LIBRARIES & CONNECTIONS
try:
    from supabase import create_client, Client
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Border, Side, Alignment, Font
except ImportError:
    st.error("⚠️ Thiếu thư viện. Vui lòng chạy lệnh: pip install streamlit pandas supabase google-api-python-client google-auth-oauthlib openpyxl altair")
    st.stop()

# CONNECT SERVER
try:
    if "supabase" not in st.secrets or "google_oauth" not in st.secrets:
        st.error("⚠️ Chưa cấu hình secrets.toml. Vui lòng kiểm tra lại file secrets.")
        st.stop()

    SUPABASE_URL = st.secrets["supabase"]["url"]
    SUPABASE_KEY = st.secrets["supabase"]["key"]
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    OAUTH_INFO = st.secrets["google_oauth"]
    ROOT_FOLDER_ID = OAUTH_INFO.get("root_folder_id", "1GLhnSK7Bz7LbTC-Q7aPt_Itmutni5Rqa")
except Exception as e:
    st.error(f"⚠️ Lỗi Config: {e}"); st.stop()

# =============================================================================
# 2. HÀM HỖ TRỢ (UTILS)
# =============================================================================
import requests

# --- CẤU HÌNH TELEGRAM ---
TELEGRAM_BOT_TOKEN = "7785342410:AAHcdXRCu6qZs-M4mGowF-65AAGzc1kdXjw"
TELEGRAM_GROUP_ID = "-5283852302"  # <--- BẠN THAY ID GROUP VỪA LẤY VÀO ĐÂY (Nhớ giữ nguyên dấu trừ)

def send_telegram_notification(assignee_name, issue_desc, new_status, new_progress):
    import streamlit as st
    import requests
    
    # Mặc định gửi tất cả thông báo vào Group ID
    chat_id = TELEGRAM_GROUP_ID
    
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    message = (
        f"🔔 <b>CẬP NHẬT TIẾN ĐỘ SỰ CỐ</b>\n\n"
        f"👤 <b>Phụ trách:</b> {assignee_name}\n"
        f"📝 <b>Vấn đề:</b> {issue_desc}\n"
        f"📊 <b>Tiến độ mới:</b> {new_progress}\n"
        f"🏷 <b>Trạng thái:</b> {new_status}\n\n"
        f"<i>Vui lòng kiểm tra lại phần mềm CRM để xem chi tiết!</i>"
    )
    payload = {"chat_id": chat_id, "text": message, "parse_mode": "HTML"}
    try:
        response = requests.post(url, json=payload)
        # Bắt bệnh nếu có lỗi
        if response.status_code != 200:
            st.error(f"🛑 Telegram báo lỗi: {response.text}")
        else:
            st.toast("✅ Đã gửi thông báo vào Group Team!", icon="🚀")
    except Exception as e:
        st.error(f"🛑 Lỗi hệ thống khi gửi Telegram: {e}")
# -------------------------
def get_drive_service():
    try:
        cred_info = OAUTH_INFO
        creds = Credentials(None, refresh_token=cred_info["refresh_token"], 
                            token_uri="https://oauth2.googleapis.com/token", 
                            client_id=cred_info["client_id"], client_secret=cred_info["client_secret"])
        return build('drive', 'v3', credentials=creds)
    except: return None

# Hàm tạo folder đệ quy
def get_or_create_folder_hierarchy(srv, path_list, parent_id):
    current_parent_id = parent_id
    for folder_name in path_list:
        q = f"'{current_parent_id}' in parents and name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        results = srv.files().list(q=q, fields="files(id)").execute().get('files', [])
        
        if results:
            current_parent_id = results[0]['id']
        else:
            file_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [current_parent_id]
            }
            folder = srv.files().create(body=file_metadata, fields='id').execute()
            current_parent_id = folder.get('id')
            try: srv.permissions().create(fileId=current_parent_id, body={'role': 'reader', 'type': 'anyone'}).execute()
            except: pass
            
    return current_parent_id

def upload_to_drive_structured(file_obj, path_list, file_name):
    srv = get_drive_service()
    if not srv: return "", ""
    try:
        folder_id = get_or_create_folder_hierarchy(srv, path_list, ROOT_FOLDER_ID)
        media = MediaIoBaseUpload(file_obj, mimetype=mimetypes.guess_type(file_name)[0] or 'application/octet-stream', resumable=True)
        file_meta = {'name': file_name, 'parents': [folder_id]}
        q_ex = f"'{folder_id}' in parents and name='{file_name}' and trashed=false"
        exists = srv.files().list(q=q_ex, fields="files(id)").execute().get('files', [])
        if exists:
            file_id = exists[0]['id']
            srv.files().update(fileId=file_id, media_body=media).execute()
        else:
            file_id = srv.files().create(body=file_meta, media_body=media, fields='id').execute()['id']
        try: srv.permissions().create(fileId=file_id, body={'role': 'reader', 'type': 'anyone'}).execute()
        except: pass
        folder_link = f"https://drive.google.com/drive/folders/{folder_id}"
        return folder_link, file_id
    except Exception as e: 
        st.error(f"Lỗi upload Drive: {e}")
        return "", ""

def upload_to_drive_simple(file_obj, sub_folder, file_name):
    srv = get_drive_service()
    if not srv: return "", ""
    try:
        folder_id = get_or_create_folder_hierarchy(srv, [sub_folder], ROOT_FOLDER_ID)
        media = MediaIoBaseUpload(file_obj, mimetype=mimetypes.guess_type(file_name)[0] or 'application/octet-stream', resumable=True)
        file_meta = {'name': file_name, 'parents': [folder_id]}
        q_ex = f"'{folder_id}' in parents and name='{file_name}' and trashed=false"
        exists = srv.files().list(q=q_ex, fields="files(id)").execute().get('files', [])
        if exists:
            file_id = exists[0]['id']
            srv.files().update(fileId=file_id, media_body=media).execute()
        else:
            file_id = srv.files().create(body=file_meta, media_body=media, fields='id').execute()['id']
        try: srv.permissions().create(fileId=file_id, body={'role': 'reader', 'type': 'anyone'}).execute()
        except: pass
        return f"https://drive.google.com/thumbnail?id={file_id}&sz=w200", file_id
    except: return "", ""

def search_file_in_drive_by_name(name_contains):
    srv = get_drive_service()
    if not srv: return None, None, None
    try:
        q = f"name contains '{name_contains}' and trashed=false"
        results = srv.files().list(q=q, fields="files(id, name, parents)").execute().get('files', [])
        if results:
            return results[0]['id'], results[0]['name'], (results[0]['parents'][0] if 'parents' in results[0] else None)
        return None, None, None
    except: return None, None, None

def download_from_drive(file_id):
    srv = get_drive_service()
    if not srv: return None
    try:
        request = srv.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False: status, done = downloader.next_chunk()
        
        # --- FIX QUAN TRỌNG: Đưa con trỏ về đầu file để pandas đọc được ---
        fh.seek(0) 
        return fh
    except: return None

def safe_str(val):
    if val is None: return ""
    s = str(val).strip()
    if s.lower() in ['nan', 'none', 'null', 'nat', '']: return ""
    return s

def to_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).replace(",", "").replace("¥", "").replace("$", "").replace("RMB", "").replace("VND", "").replace(" ", "").upper()
    try:
        nums = re.findall(r"[-+]?\d*\.\d+|\d+", s)
        return float(nums[0]) if nums else 0.0
    except: return 0.0

def fmt_num(x): 
    try:
        if x is None: return "0"
        val = float(x)
        if val.is_integer(): return "{:,.0f}".format(val)
        else:
            s = "{:,.3f}".format(val)
            return s.rstrip('0').rstrip('.')
    except: return "0"

# --- NEW: FORMAT 2 DECIMAL PLACES (FOR QUOTE TAB) ---
def fmt_float_2(x):
    try:
        if x is None: return "0.00"
        val = float(x)
        return "{:,.2f}".format(val)
    except: return "0.00"

def clean_key(s): return safe_str(s).lower()

def calc_eta(order_date_str, leadtime_val):
    try:
        if isinstance(order_date_str, datetime): dt_order = order_date_str
        else:
            try: dt_order = datetime.strptime(order_date_str, "%d/%m/%Y")
            except: dt_order = datetime.now()
        lt_str = str(leadtime_val)
        nums = re.findall(r'\d+', lt_str)
        days = int(nums[0]) if nums else 0
        dt_exp = dt_order + timedelta(days=days)
        return dt_exp.strftime("%d/%m/%Y")
    except: return ""

def load_data(table, order_by="id", ascending=True):
    try:
        query = supabase.table(table).select("*")
        res = query.execute()
        df = pd.DataFrame(res.data)
        
        if not df.empty:
           # Giữ lại cột 'id' cho crm_issues để phục vụ việc Edit/Xóa
            if table not in ["crm_tracking", "crm_payments", "crm_issues"] and 'id' in df.columns: 
                df = df.drop(columns=['id'])
            
            # Sort bằng Pandas nếu có cột order_by
            if order_by in df.columns:
                df = df.sort_values(by=order_by, ascending=ascending)
            
        return df
    except Exception as e:
        # st.error(f"Lỗi load data {table}: {e}") # Có thể uncomment để debug
        return pd.DataFrame()

# =============================================================================
# 3. LOGIC TÍNH TOÁN CORE (UPDATED: MANUAL OVERRIDE SUPPORT & NEW PROFIT FORMULA)
# =============================================================================
def recalculate_quote_logic(df, params):
    # 1. Chuyển đổi dữ liệu sang số (Float) để tính toán
    cols_money_input = [
        "Q'ty", "Buying price(VND)", "Buying price(RMB)", "Exchange rate",
        "AP price(VND)", "Unit price(VND)", 
        "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", 
        "Transportation", "Management fee(%)", "Payback(%)"
    ]
    
    # Tạo cột nếu chưa có (để tránh lỗi) và chuyển sang số
    for c in cols_money_input:
        if c not in df.columns: df[c] = 0.0
        df[c] = df[c].apply(to_float)

    # 2. TÍNH TOÁN CÁC CỘT TOTAL & LOGIC CƠ BẢN (Luôn chạy)
    # Buying VND luôn = RMB * Rate 
    df["Buying price(VND)"] = df["Buying price(RMB)"] * df["Exchange rate"]
    
    df["Total buying price(VND)"] = df["Buying price(VND)"] * df["Q'ty"]
    df["Total buying price(rmb)"] = df["Buying price(RMB)"] * df["Q'ty"]
    df["AP total price(VND)"] = df["AP price(VND)"] * df["Q'ty"]
    df["Total price(VND)"] = df["Unit price(VND)"] * df["Q'ty"]
    
    # GAP là kết quả tính toán
    df["GAP"] = df["Total price(VND)"] - df["AP total price(VND)"]

    # 3. TÍNH LỢI NHUẬN (PROFIT)
    # --- UPDATED FORMULA ---
    # Profit = Total price - (Total buying price VND + GAP + End user + Buyer + Import tax + VAT + Transportation + Management fee) + Payback
    
    # Lưu ý: GAP trong công thức này là giá trị GAP thô (Total - AP Total) như yêu cầu.
    
    # Cộng dồn các chi phí (bao gồm GAP)
    cost_ops = (df["Total buying price(VND)"] + 
                df["GAP"] +
                df["End user(%)"] + 
                df["Buyer(%)"] + 
                df["Import tax(%)"] + 
                df["VAT"] + 
                df["Transportation"] + 
                df["Management fee(%)"])

    # Lợi nhuận = Doanh thu - Chi phí + Payback
    df["Profit(VND)"] = df["Total price(VND)"] - cost_ops + df["Payback(%)"]
    
    # Tính % Lợi nhuận
    df["Profit_Pct_Raw"] = df.apply(lambda row: (row["Profit(VND)"] / row["Total price(VND)"] * 100) if row["Total price(VND)"] > 0 else 0, axis=1)
    df["Profit(%)"] = df["Profit_Pct_Raw"].apply(lambda x: f"{x:.1f}%")
    
    # Cảnh báo
    def set_warning(row):
        if "KHÔNG KHỚP" in str(row["Cảnh báo"]): return row["Cảnh báo"]
        return "⚠️ LOW" if row["Profit_Pct_Raw"] < 10 else "✅ OK"
    df["Cảnh báo"] = df.apply(set_warning, axis=1)

    return df

# --- IMPROVED FORMULA PARSER ---
def parse_formula(formula, buying_price, ap_price):
    if not formula: return 0.0
    
    # 1. Normalize: Uppercase and Strip
    s = str(formula).strip().upper()
    
    # 2. Handle '='
    if s.startswith("="): s = s[1:]
    
    # 3. Replace Keywords (Longer first to avoid substrings issue)
    # Handle 'AP PRICE' explicitly before 'AP'
    s = s.replace("AP PRICE", str(ap_price))
    s = s.replace("BUYING PRICE", str(buying_price))
    
    # Handle shorthands
    s = s.replace("AP", str(ap_price))
    s = s.replace("BUY", str(buying_price))
    
    # 4. Cleanup Syntax
    s = s.replace(",", ".").replace("%", "/100").replace("X", "*")
    
    # 5. Filter Unsafe Characters (Only digits, dots, math ops)
    s = re.sub(r'[^0-9.+\-*/()]', '', s)
    
    try: 
        if not s: return 0.0
        return float(eval(s))
    except: return 0.0

# =============================================================================
# 4. GIAO DIỆN CHÍNH
# =============================================================================
t1, t2, t3, t4, t5, t7, t8, t6 = st.tabs(["📊 DASHBOARD", "📦 KHO HÀNG", "💰 BÁO GIÁ", "📑 QUẢN LÝ PO", "🚚 TRACKING", "🚀 DỰ ÁN", "⚠️ QUẢN LÝ ISSUE", "⚙️ MASTER DATA"])
# --- TAB 1: DASHBOARD (UPDATED - FIX METRICS LOGIC) ---
# =============================================================================
with t1:
    # --- 1. HEADER & ADMIN RESET ---
    c_h1, c_h2 = st.columns([3, 1])
    with c_h1:
        if st.button("🔄 REFRESH DATA"): st.cache_data.clear(); st.rerun()
    
    with c_h2:
        with st.popover("⚠️ RESET SYSTEM"):
            st.markdown("**Xóa dữ liệu giao dịch (Giữ lại Khách/NCC/Kho)**")
            adm_pass_reset = st.text_input("Mật khẩu Admin", type="password", key="pass_reset_db")
            if st.button("🔴 XÓA SẠCH LỊCH SỬ"):
                if adm_pass_reset == "admin":
                    try:
                        # Xóa các bảng Transaction (Lịch sử, PO, Tracking, Payment)
                        supabase.table("crm_shared_history").delete().neq("id", 0).execute()
                        supabase.table("db_customer_orders").delete().neq("id", 0).execute()
                        supabase.table("db_supplier_orders").delete().neq("id", 0).execute()
                        supabase.table("crm_tracking").delete().neq("id", 0).execute()
                        supabase.table("crm_payments").delete().neq("id", 0).execute()
                        
                        st.toast("✅ Đã reset toàn bộ hệ thống về trạng thái ban đầu!", icon="🗑️")
                        time.sleep(1.5)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Lỗi khi xóa: {e}")
                else:
                    st.error("Sai mật khẩu!")

    # --- 2. LOAD DATA ---
    db_cust_po = load_data("db_customer_orders") # Nguồn PO Khách hàng (Tham khảo)
    db_hist = load_data("crm_shared_history")    # Nguồn Lịch sử (Chính xác cho Doanh thu/Chi phí/Lợi nhuận)
    db_items = load_data("crm_purchases")        # Master Data

    # --- 3. METRICS CALCULATION (FIXED) ---
    # Logic cũ sai vì lấy Doanh thu từ PO nhưng Chi phí từ History.
    # Logic mới: Lấy TẤT CẢ từ History để đảm bảo (Doanh thu - Chi phí = Lợi nhuận)
    
    revenue_total = 0
    profit_total = 0
    cost_total = 0
    total_po_raw = db_cust_po['total_price'].apply(to_float).sum() if not db_cust_po.empty else 0

    if not db_hist.empty:
        # Lấy Doanh thu từ những đơn ĐÃ CÓ lịch sử chi phí
        revenue_total = db_hist['total_price_vnd'].apply(to_float).sum()
        # Lấy Lợi nhuận thực tế
        profit_total = db_hist['profit_vnd'].apply(to_float).sum()
        # Tính chi phí khớp với doanh thu này
        cost_total = revenue_total - profit_total
    else:
        # Nếu chưa có history thì hiển thị doanh thu thô, lợi nhuận = 0
        revenue_total = total_po_raw
    
    # --- 4. KPI CARDS ---
    c1, c2, c3 = st.columns(3)
    
    # Hiển thị Note nhỏ để biết tổng PO thực tế nếu lệch với History
    delta_msg = ""
    if total_po_raw > revenue_total:
        delta_msg = f" (Tổng PO thực: {fmt_num(total_po_raw)})"

    c1.markdown(f"<div class='card-3d bg-sales'><h3>DOANH THU (Đã chốt Cost)</h3><h1>{fmt_num(revenue_total)}</h1><p style='font-size:12px; margin:0;'>{delta_msg}</p></div>", unsafe_allow_html=True)
    c2.markdown(f"<div class='card-3d bg-cost'><h3>CHI PHÍ (Formula)</h3><h1>{fmt_num(cost_total)}</h1></div>", unsafe_allow_html=True)
    c3.markdown(f"<div class='card-3d bg-profit'><h3>LỢI NHUẬN (Est.)</h3><h1>{fmt_num(profit_total)}</h1></div>", unsafe_allow_html=True)

    st.divider()

    # --- 5. CHARTS ---
    if not db_hist.empty:
        # Pre-process Data
        db_hist['date_dt'] = pd.to_datetime(db_hist['date'], format="%Y-%m-%d", errors='coerce')
        db_hist['Month'] = db_hist['date_dt'].dt.strftime('%Y-%m')
        
        # Map Type
        type_map = {}
        if not db_items.empty:
            for r in db_items.to_dict('records'):
                type_map[clean_key(r.get('item_code'))] = safe_str(r.get('type', 'Other'))
        
        db_hist['Type'] = db_hist['item_code'].apply(lambda x: type_map.get(clean_key(x), "Other"))
        db_hist['Revenue'] = db_hist['total_price_vnd'].apply(to_float)
        
        # -----------------------------------------------------------
        # CHART 1: CỘT & TREND (DOANH SỐ THEO THÁNG & KHÁCH HÀNG)
        # -----------------------------------------------------------
        st.subheader("📈 Xu hướng Doanh số & Khách hàng")
        
        # Group Data
        chart_data = db_hist.groupby(['Month', 'customer'])['Revenue'].sum().reset_index()
        
        # Base Chart
        base = alt.Chart(chart_data).encode(x=alt.X('Month', title='Tháng'))
        
        # Bar Chart
        bar = base.mark_bar().encode(
            y=alt.Y('Revenue', title='Doanh thu (VND)'),
            color=alt.Color('customer', title='Khách hàng'),
            tooltip=['Month', 'customer', alt.Tooltip('Revenue', format=',.0f')]
        )
        
        # Text Labels for Bar
        text_bar = base.mark_text(dy=3, color='white').encode(
            y=alt.Y('Revenue', stack='zero'),
            text=alt.Text('Revenue', format='.2s') 
        )

        # Trend Line (Total per Month)
        line_data = db_hist.groupby(['Month'])['Revenue'].sum().reset_index()
        base_line = alt.Chart(line_data).encode(x='Month')
        
        line = base_line.mark_line(color='red', point=True).encode(
            y='Revenue',
            tooltip=[alt.Tooltip('Revenue', format=',.0f', title='Tổng Trend')]
        )
        
        text_line = base_line.mark_text(align='center', baseline='bottom', dy=-10, color='red').encode(
            y='Revenue',
            text=alt.Text('Revenue', format=',.0f')
        )
        
        st.altair_chart((bar + text_bar + line + text_line).interactive(), use_container_width=True)
        
        # -----------------------------------------------------------
        # CHART 2 & 3: PIE CHARTS
        # -----------------------------------------------------------
        st.divider()
        st.subheader("🍰 Cơ cấu Doanh số")
        col_pie1, col_pie2 = st.columns(2)
        
        def create_pie_chart_with_labels(df_source, group_col, value_col, color_scheme="category20"):
            df_agg = df_source.groupby(group_col)[value_col].sum().reset_index()
            total_val = df_agg[value_col].sum()
            df_agg['Percent'] = (df_agg[value_col] / total_val * 100).round(1)
            df_agg['Label'] = df_agg.apply(lambda x: f"{x['Percent']}% ({fmt_num(x[value_col])})", axis=1)
            
            base = alt.Chart(df_agg).encode(theta=alt.Theta(field=value_col, type="quantitative", stack=True))
            pie = base.mark_arc(outerRadius=120).encode(
                color=alt.Color(field=group_col, type="nominal", scale=alt.Scale(scheme=color_scheme)),
                order=alt.Order(field=value_col, sort="descending"),
                tooltip=[group_col, alt.Tooltip(value_col, format=',.0f'), 'Percent']
            )
            text = base.mark_text(radius=140).encode(
                text=alt.Text("Label"),
                order=alt.Order(field=value_col, sort="descending"),
                color=alt.value("black") 
            )
            return (pie + text)

        with col_pie1:
            st.write("**Theo Khách Hàng**")
            chart_pie_cust = create_pie_chart_with_labels(db_hist, 'customer', 'Revenue', 'tableau10')
            st.altair_chart(chart_pie_cust, use_container_width=True)
            
        with col_pie2:
            st.write("**Theo Loại Sản Phẩm (Type)**")
            chart_pie_type = create_pie_chart_with_labels(db_hist, 'Type', 'Revenue', 'set2')
            st.altair_chart(chart_pie_type, use_container_width=True)
            
    else:
        st.info("Chưa có dữ liệu lịch sử để vẽ biểu đồ. Hãy tạo Báo Giá và Lưu Lịch Sử.")

# --- TAB 2: KHO HÀNG (FINAL FIX: FORCE TYPE CASTING) ---
with t2:
    st.subheader("QUẢN LÝ KHO HÀNG (Excel Online)")
    c_imp, c_view = st.columns([0.5, 5])
    
    # --- HÀM LÀM SẠCH DỮ LIỆU ---
    def clean_strict(val):
        if val is None: return ""
        return re.sub(r'\s+', '', str(val)).lower()

    # --- CỘT TRÁI: IMPORT (GIỮ NGUYÊN) ---
    with c_imp:
        st.markdown("**📥 Import Kho Hàng**")
        st.caption("Excel cột A->O")
        st.info("No, Code, Name, Specs, Qty, BuyRMB, TotalRMB, Rate, BuyVND, TotalVND, Leadtime, Supplier, Images, Type, N/U/O/C")
        
        with st.expander("🛠️ Reset DB"):
            adm_pass = st.text_input("Pass", type="password", key="adm_inv")
            if st.button("⚠️ XÓA SẠCH"):
                if adm_pass == "admin":
                    supabase.table("crm_purchases").delete().neq("id", 0).execute()
                    st.success("Deleted!"); time.sleep(1); st.rerun()
                else: st.error("Sai Pass!")
        
        up_file = st.file_uploader("Upload Excel", type=["xlsx"], key="inv_up")
            
        if up_file and st.button("🚀 Kiểm tra & Import"):
            try:
                wb = load_workbook(up_file, data_only=False); ws = wb.active
                
                img_map = {}
                detected_images = []
                for image in getattr(ws, '_images', []):
                    try:
                        r_idx = image.anchor._from.row + 1
                        cell_specs = ws.cell(row=r_idx, column=4).value 
                        specs_val = safe_str(cell_specs)
                        safe_name = re.sub(r'[\\/*?:"<>|]', "", specs_val).strip()
                        if not safe_name: safe_name = f"NO_SPECS_R{r_idx}"
                        fname = f"{safe_name}.png"
                        detected_images.append({'row': r_idx, 'name': fname, 'data': image._data()})
                    except: continue

                detected_images.sort(key=lambda x: x['row'])

                for img in detected_images:
                    r = img['row']
                    buf = io.BytesIO(img['data'])
                    link, _ = upload_to_drive_simple(buf, "CRM_PRODUCT_IMAGES", img['name'])
                    if r not in img_map: img_map[r] = link
                    elif (r + 1) not in img_map: img_map[r + 1] = link

                df = pd.read_excel(up_file, header=None, skiprows=1, dtype=str).fillna("")
                raw_records = []
                prog = st.progress(0)
                cols_map = ["no", "item_code", "item_name", "specs", "qty", "buying_price_rmb", 
                            "total_buying_price_rmb", "exchange_rate", "buying_price_vnd", 
                            "total_buying_price_vnd", "leadtime", "supplier_name", "image_path", "type", "nuoc"]

                for i, r in df.iterrows():
                    d = {}
                    for idx, field in enumerate(cols_map):
                        if idx < len(r): d[field] = safe_str(r.iloc[idx])
                        else: d[field] = ""
                    
                    has_data = d['item_code'] or d['item_name'] or d['specs']
                    if has_data:
                        if not d.get('image_path') and (i+2) in img_map: 
                            d['image_path'] = img_map[i+2]
                            
                        d['row_order'] = i + 1 
                        d['qty'] = to_float(d.get('qty', 0))
                        d['buying_price_rmb'] = to_float(d['buying_price_rmb'])
                        d['total_buying_price_rmb'] = to_float(d['total_buying_price_rmb'])
                        d['exchange_rate'] = to_float(d['exchange_rate'])
                        d['buying_price_vnd'] = to_float(d['buying_price_vnd'])
                        d['total_buying_price_vnd'] = to_float(d['total_buying_price_vnd'])
                        raw_records.append(d)
                    prog.progress((i + 1) / len(df))
                
                if raw_records:
                    best_records_map = {} 
                    for rec in raw_records:
                        sig = (clean_strict(rec.get('item_code')), clean_strict(rec.get('item_name')), 
                               clean_strict(rec.get('specs')), clean_strict(rec.get('nuoc')))
                        price_curr = rec['buying_price_rmb']
                        
                        if sig not in best_records_map:
                            best_records_map[sig] = rec
                        else:
                            price_exist = best_records_map[sig]['buying_price_rmb']
                            if price_curr < price_exist: best_records_map[sig] = rec
                    
                    processed_records = list(best_records_map.values())

                    df_db = load_data("crm_purchases")
                    existing_sigs = set()
                    if not df_db.empty:
                        for r in df_db.to_dict('records'):
                            sig_db = (clean_strict(r.get('item_code')), clean_strict(r.get('item_name')), 
                                      clean_strict(r.get('specs')), clean_strict(r.get('nuoc')))
                            existing_sigs.add(sig_db)
                    
                    dups = []
                    non_dups = []
                    for rec in processed_records:
                        sig_rec = (clean_strict(rec.get('item_code')), clean_strict(rec.get('item_name')), 
                                   clean_strict(rec.get('specs')), clean_strict(rec.get('nuoc')))
                        if sig_rec in existing_sigs: dups.append(rec)
                        else: non_dups.append(rec)
                    
                    st.session_state.import_dups = dups
                    st.session_state.import_non_dups = non_dups
                    st.session_state.import_step = "confirm" if dups else "auto_import"
                    st.rerun()

            except Exception as e: st.error(f"Lỗi xử lý file: {e}")

        step = st.session_state.get("import_step", None)
        if step == "confirm":
            st.warning(f"⚠️ Có {len(st.session_state.import_dups)} item trùng lặp.")
            with st.expander("Xem chi tiết"):
                st.dataframe(pd.DataFrame(st.session_state.import_dups)[['item_code', 'item_name', 'specs']], hide_index=True)
            
            c1, c2 = st.columns(2)
            if c1.button("✅ Chỉ Import dòng mới"):
                st.session_state.final_import_list = st.session_state.import_non_dups
                st.session_state.import_step = "executing"; st.rerun()
            if c2.button("⚠️ Import TẤT CẢ"):
                st.session_state.final_import_list = st.session_state.import_dups + st.session_state.import_non_dups
                st.session_state.import_step = "executing"; st.rerun()

        elif step == "auto_import":
            st.session_state.final_import_list = st.session_state.import_non_dups
            st.session_state.import_step = "executing"; st.rerun()

        elif step == "executing":
            final_list = st.session_state.get("final_import_list", [])
            if final_list:
                try:
                    chunk_ins = 100
                    for k in range(0, len(final_list), chunk_ins):
                        batch = final_list[k:k+chunk_ins]
                        try: supabase.table("crm_purchases").insert(batch).execute()
                        except Exception as e_ins:
                             if "row_order" in str(e_ins):
                                for rec in batch: 
                                    if 'row_order' in rec: del rec['row_order']
                                supabase.table("crm_purchases").insert(batch).execute()
                             else: raise e_ins
                    st.success(f"✅ Đã import {len(final_list)} dòng!"); time.sleep(1)
                    st.session_state.import_step = None; st.cache_data.clear(); st.rerun()
                except Exception as e:
                    st.error(f"Lỗi SQL: {e}"); st.session_state.import_step = None

    # --- CỘT PHẢI: HIỂN THỊ (SỬA LỖI STREAMLIT API EXCEPTION) ---
    with c_view:
        try:
            # Load Data và giữ ID
            res = supabase.table("crm_purchases").select("*").execute()
            df_pur = pd.DataFrame(res.data)
            if not df_pur.empty and 'row_order' in df_pur.columns:
                df_pur = df_pur.sort_values(by='row_order', ascending=True)
        except:
            df_pur = pd.DataFrame()

        # Drop cột rác
        cols_to_drop = ['created_at', 'row_order']
        df_pur = df_pur.drop(columns=[c for c in cols_to_drop if c in df_pur.columns], errors='ignore')

        search = st.text_input("🔍 Tìm kiếm (Name, Code, Specs...)", key="search_pur")
        
        if not df_pur.empty:
            if search:
                mask = df_pur.astype(str).apply(lambda x: x.str.contains(search, case=False, na=False)).any(axis=1)
                df_pur = df_pur[mask]
            
            # --- 1. CHUẨN HÓA DỮ LIỆU (FIX TRIỆT ĐỂ LỖI TYPE) ---
            # Thêm cột Select (Boolean)
            df_pur.insert(0, "Select", False)
            
            # Ép kiểu cột Qty (Float/Int)
            if 'qty' in df_pur.columns:
                df_pur['qty'] = pd.to_numeric(df_pur['qty'], errors='coerce').fillna(0)
                
            # Định dạng các cột tiền (Về String)
            cols_money = ["buying_price_vnd", "total_buying_price_vnd", "buying_price_rmb", "total_buying_price_rmb"]
            for c in cols_money:
                if c in df_pur.columns: df_pur[c] = df_pur[c].apply(fmt_num)

            # ÉP TẤT CẢ CÁC CỘT CÒN LẠI VỀ STRING (Trừ Select, qty, id)
            # Điều này ngăn chặn việc Streamlit hiểu nhầm None/NaN là float trong cột Text
            exclude_cols = ['Select', 'qty', 'id']
            for col in df_pur.columns:
                if col not in exclude_cols:
                    df_pur[col] = df_pur[col].fillna("").astype(str)

            # --- 2. SẮP XẾP CỘT (NO -> VỊ TRÍ 2) ---
            # Thứ tự mong muốn: [Select, no, item_code, ..., id]
            cols = df_pur.columns.tolist()
            if 'no' in cols:
                cols.remove('no')
                cols.insert(1, 'no') # Vị trí index 1 (sau Select ở 0)
                df_pur = df_pur[cols]

            # --- 3. CẤU HÌNH HIỂN THỊ ---
            column_config = {
                "Select": st.column_config.CheckboxColumn("Chọn", width="small"),
                "no": st.column_config.TextColumn("No.", width="small"),
                "id": None, # Ẩn cột ID (Dùng None là chuẩn nhất)
                "image_path": st.column_config.ImageColumn("Images", width="small"),
                "item_code": st.column_config.TextColumn("Code", width="medium"),
                "item_name": st.column_config.TextColumn("Name", width="medium"),
                "specs": st.column_config.TextColumn("Specs", width="large"),
                "nuoc": st.column_config.TextColumn("N/U/O/C", width="small"),
                "buying_price_vnd": st.column_config.TextColumn("Buying (VND)"),
                "qty": st.column_config.NumberColumn("Qty", format="%d"),
            }

            edited_df = st.data_editor(
                df_pur,
                column_config=column_config,
                use_container_width=True,
                height=700,
                hide_index=True,
                key="data_editor_inventory"
            )

            # --- XỬ LÝ XÓA DÒNG ---
            selected_rows = edited_df[edited_df["Select"] == True]
            if not selected_rows.empty:
                st.divider()
                st.warning(f"🛑 Đang chọn xóa {len(selected_rows)} dòng.")
                c_del1, c_del2 = st.columns([2, 1])
                pass_del = c_del1.text_input("Mật khẩu Admin để xóa:", type="password", key="pass_del_row")
                
                if c_del2.button("🔥 XÁC NHẬN XÓA"):
                    if pass_del == "admin":
                        try:
                            # Lấy ID từ dataframe
                            ids_to_delete = selected_rows['id'].tolist()
                            if ids_to_delete:
                                supabase.table("crm_purchases").delete().in_("id", ids_to_delete).execute()
                                st.success(f"Đã xóa {len(ids_to_delete)} dòng!"); time.sleep(1)
                                st.cache_data.clear(); st.rerun()
                            else: st.error("Lỗi: Không tìm thấy ID.")
                        except Exception as e: st.error(f"Lỗi xóa: {e}")
                    else: st.error("Sai mật khẩu!")
        else: st.info("Kho hàng trống.")
import re
import json
import time
from datetime import datetime
import pandas as pd
import numpy as np
import io
from openpyxl import load_workbook, Workbook

import re
import json
import time
from datetime import datetime
import pandas as pd
import numpy as np
import io
from openpyxl import load_workbook, Workbook

import re
import json
import time
from datetime import datetime
import pandas as pd
import numpy as np
import io
from openpyxl import load_workbook, Workbook

import re
import json
import time
from datetime import datetime
import pandas as pd
import numpy as np
import io
from openpyxl import load_workbook, Workbook

import re
import json
import time
from datetime import datetime
import pandas as pd
import numpy as np
import io
from openpyxl import load_workbook, Workbook

# =============================================================================
# --- TAB 3: BÁO GIÁ (GIỮ NGUYÊN NHƯ CŨ - CHỈ CHỈNH LẠI FORMAT CHUẨN) ---
# =============================================================================
with t3:
    # --- A. CÁC HÀM HỖ TRỢ NỘI BỘ ---
    def local_parse_money(val):
        try:
            if pd.isna(val) or str(val).strip() == "": return 0.0
            # Xử lý an toàn cho cả dấu phẩy và chấm
            s = str(val).replace(",", "").replace("%", "").strip()
            return float(s)
        except: return 0.0

    def local_fmt_vnd(val):
        try:
            if pd.isna(val): return "0"
            return "{:,.0f}".format(round(float(val)))
        except: return str(val)

    def local_fmt_rmb(val):
        try:
            if pd.isna(val): return "0.00"
            return "{:,.2f}".format(float(val))
        except: return str(val)
    
    def to_float(val): return local_parse_money(val)

    def clean_number_for_db(val):
        try:
            f = float(val)
            if f.is_integer(): return int(f) 
            return f
        except: return val

    def normalize_match_str(val):
        if pd.isna(val) or val is None: return ""
        s = str(val).lower().strip()
        s = re.sub(r'[^a-z0-9]', '', s) 
        return s

    def local_eval_formula(formula_str, val_buy, val_ap):
        if not formula_str: return 0.0
        s = str(formula_str).strip().upper()
        if s.startswith("="): s = s[1:]
        s = s.replace("AP PRICE", str(val_ap)).replace("BUYING PRICE", str(val_buy))
        s = s.replace("AP", str(val_ap)).replace("BUY", str(val_buy))
        s = s.replace(",", ".").replace("%", "/100").replace("X", "*")
        s = re.sub(r'[^0-9.+\-*/()]', '', s)
        try: 
            if not s: return 0.0
            return float(eval(s))
        except: return 0.0

    # --- B. HÀM TÍNH TOÁN LOGIC TAB 3 ---
    def recalculate_quote_logic(df, params):
        if df.empty: return df
        p_end = params.get('end', 0) / 100.0
        p_buy = params.get('buy', 0) / 100.0
        p_tax = params.get('tax', 0) / 100.0
        p_vat = params.get('vat', 0) / 100.0
        p_mgmt = params.get('mgmt', 0) / 100.0
        p_pay = params.get('pay', 0) / 100.0
        v_trans = float(params.get('trans', 0))

        for idx, row in df.iterrows():
            try:
                qty = local_parse_money(row.get("Q'ty", 0))
                buy_rmb = local_parse_money(row.get("Buying price(RMB)", 0))
                ex_rate = local_parse_money(row.get("Exchange rate", 0))
                
                buy_vnd_unit = local_parse_money(row.get("Buying price(VND)", 0))
                if buy_rmb > 0 and ex_rate > 0:
                    buy_vnd_unit = round(buy_rmb * ex_rate, 0)

                total_buy_vnd = round(buy_vnd_unit * qty, 0)
                total_buy_rmb = round(buy_rmb * qty, 2)
                
                ap_vnd_unit = local_parse_money(row.get("AP price(VND)", 0))
                ap_total = round(ap_vnd_unit * qty, 0)
                
                unit_price = local_parse_money(row.get("Unit price(VND)", 0))
                total_price = round(unit_price * qty, 0)
                
                gap = total_price - ap_total

                val_imp_tax = round(total_buy_vnd * p_tax, 0)
                val_end = round(ap_total * p_end, 0) 
                val_buyer = round(total_price * p_buy, 0)
                val_vat = round(total_price * p_vat, 0)
                val_mgmt = round(total_price * p_mgmt, 0)
                val_trans = round(v_trans, 0)
                
                val_payback = round(gap * p_pay, 0) if gap > 0 else 0.0

                sum_deductions = (total_buy_vnd + gap + val_end + val_buyer + val_imp_tax + val_vat + val_trans + val_mgmt)
                val_profit = round(total_price - sum_deductions + val_payback, 0)
                
                pct_profit = 0.0
                if total_price != 0: pct_profit = (val_profit / total_price) * 100

                current_warning = str(row.get("Cảnh báo", "")).replace("⚠️ (<10%)", "").replace("|", "").strip()
                if pct_profit < 10.0:
                    if current_warning: current_warning += " | "
                    current_warning += "⚠️ (<10%)"
                
                df.at[idx, "Cảnh báo"] = current_warning
                df.at[idx, "Buying price(VND)"] = buy_vnd_unit
                df.at[idx, "Total buying price(rmb)"] = total_buy_rmb
                df.at[idx, "Total buying price(VND)"] = total_buy_vnd
                df.at[idx, "AP total price(VND)"] = ap_total
                df.at[idx, "Total price(VND)"] = total_price
                df.at[idx, "GAP"] = gap
                df.at[idx, "Import tax(%)"] = val_imp_tax
                df.at[idx, "End user(%)"] = val_end
                df.at[idx, "Buyer(%)"] = val_buyer
                df.at[idx, "VAT"] = val_vat
                df.at[idx, "Management fee(%)"] = val_mgmt
                df.at[idx, "Transportation"] = val_trans
                df.at[idx, "Payback(%)"] = val_payback
                df.at[idx, "Profit(VND)"] = val_profit
                df.at[idx, "Profit(%)"] = f"{pct_profit:.2f}%"
                
            except Exception: continue      
        return df

    if 'quote_df' not in st.session_state: st.session_state.quote_df = pd.DataFrame()
    
    # -------------------------------------------------------------------------
    # UI TAB 3
    # -------------------------------------------------------------------------
    with st.expander("🛠️ ADMIN: QUẢN LÝ LỊCH SỬ BÁO GIÁ"):
        c_adm1, c_adm2 = st.columns([3, 1])
        c_adm1.warning("⚠️ Chức năng này sẽ xóa vĩnh viễn dữ liệu.")
        adm_pass_q = c_adm2.text_input("Mật khẩu Admin", type="password", key="pass_reset_quote_tab3")
        
        if c_adm2.button("🔴 XÓA HẾT LỊCH SỬ", key="btn_clear_hist_tab3"):
            if adm_pass_q == "admin": 
                try:
                    st.cache_data.clear()
                    try: supabase.table("crm_shared_history").delete().neq("history_id", "0").execute()
                    except: pass
                    supabase.table("crm_quotations_log").delete().neq("history_id", "0").execute()
                    st.toast("✅ Đã xóa sạch 100% dữ liệu và Cache!", icon="🗑️")
                    time.sleep(1)
                    st.rerun()
                except Exception as e: st.error(f"Lỗi xóa DB: {e}")
            else: st.error("Sai mật khẩu!")

    with st.expander("🔎 TRA CỨU & TRẠNG THÁI BÁO GIÁ", expanded=False):
        c_src1, c_src2 = st.columns(2)
        search_kw = c_src1.text_input("Nhập từ khóa", help="Tìm kiếm trong lịch sử")
        up_src = c_src2.file_uploader("Hoặc Import Excel kiểm tra", type=["xlsx"], key="src_up")
        
        if st.button("Kiểm tra trạng thái"):
            df_hist = load_data("crm_quotations_log")
            df_po = load_data("db_customer_orders")
            df_items = load_data("crm_purchases") 
            item_map = {}
            if not df_items.empty:
                for r in df_items.to_dict('records'):
                    k = clean_key(r['item_code'])
                    item_map[k] = f"{safe_str(r['item_name'])} {safe_str(r['specs'])}"
            po_map = {}
            if not df_po.empty:
                for r in df_po.to_dict('records'):
                    k = f"{clean_key(r['customer'])}_{clean_key(r['item_code'])}"
                    po_map[k] = r['po_number']
            results = []
            if search_kw and not df_hist.empty:
                def check_row(row):
                    kw = search_kw.lower()
                    if kw in str(row.get('customer','')).lower(): return True
                    if kw in str(row.get('quote_no','')).lower(): return True
                    if kw in str(row.get('item_code','')).lower(): return True
                    code = clean_key(row['item_code'])
                    info = item_map.get(code, "").lower()
                    if kw in info: return True
                    return False
        
                mask = df_hist.apply(check_row, axis=1)
                found = df_hist[mask]
                for _, r in found.iterrows():
                    key = f"{clean_key(r['customer'])}_{clean_key(r['item_code'])}"
                    results.append({
                        "Trạng thái": "✅ Đã báo giá", "Customer": r['customer'], "Date": r['date'],
                        "Item Code": r['item_code'], "Info": item_map.get(clean_key(r['item_code']), ""), 
                        "Unit Price": local_fmt_vnd(r['unit_price']), "Quote No": r['quote_no'], "PO No": po_map.get(key, "---")
                    })
            if up_src:
                try:
                    df_check = pd.read_excel(up_src, dtype=str).fillna("")
                    cols_check = {clean_key(c): c for c in df_check.columns}
                    for i, r in df_check.iterrows():
                        code = ""
                        for k, col in cols_check.items():
                            if "code" in k: code = safe_str(r[col])
                        match = pd.DataFrame()
                        if not df_hist.empty and code: match = df_hist[df_hist['item_code'].str.contains(code, case=False, na=False)]
                        if not match.empty:
                            for _, m in match.iterrows():
                                key = f"{clean_key(m['customer'])}_{clean_key(m['item_code'])}"
                                results.append({
                                    "Trạng thái": "✅ Đã báo giá", "Customer": m['customer'], "Date": m['date'],
                                    "Item Code": m['item_code'], "Info": item_map.get(clean_key(m['item_code']), ""),
                                    "Unit Price": local_fmt_vnd(m['unit_price']), "Quote No": m['quote_no'], "PO No": po_map.get(key, "---")
                                })
                        else: results.append({"Trạng thái": "❌ Chưa báo giá", "Item Code": code, "Customer": "---", "Date": "---", "Unit Price": "---", "Quote No": "---", "PO No": "---"})
                except Exception as e: st.error(f"Lỗi file: {e}")
            if results: st.dataframe(pd.DataFrame(results), use_container_width=True)
            else: st.info("Không tìm thấy kết quả.")

    with st.expander("📂 XEM CHI TIẾT FILE LỊCH SỬ", expanded=False):
        df_hist_idx = load_data("crm_quotations_log", order_by="date")
        if not df_hist_idx.empty:
            df_hist_idx['display'] = df_hist_idx.apply(lambda x: f"{x['date']} | {x['customer']} | Quote: {x['quote_no']}", axis=1)
            sel_quote_hist = st.selectbox("Chọn báo giá cũ:", [""] + list(df_hist_idx['display'].unique()))
            if sel_quote_hist:
                parts = sel_quote_hist.split(" | ")
                if len(parts) >= 3:
                    q_no = parts[2].replace("Quote: ", "").strip()
                    cust = parts[1].strip()
                    hist_row = df_hist_idx[(df_hist_idx['quote_no'] == q_no) & (df_hist_idx['customer'] == cust)].iloc[0]
                    
                    config_loaded = {}
                    if 'config_data' in hist_row and hist_row['config_data']:
                        try: config_loaded = json.loads(hist_row['config_data'])
                        except: pass
                    
                    clean_config_for_ui = {}
                    if "params" in config_loaded: clean_config_for_ui = config_loaded["params"]
                    else: clean_config_for_ui = config_loaded

                    if clean_config_for_ui:
                        st.info(f"📊 **CẤU HÌNH (ĐÃ LOAD):** End:{clean_config_for_ui.get('end')}% | Buy:{clean_config_for_ui.get('buy')}%")
                        if sel_quote_hist != st.session_state.get('loaded_quote_id'):
                            for k in ["end", "buy", "tax", "vat", "pay", "mgmt", "trans"]:
                                st.session_state[f"pct_{k}"] = str(clean_config_for_ui.get(k, 0))
                                st.session_state[f"input_{k}"] = str(clean_config_for_ui.get(k, 0))
                            st.session_state.loaded_quote_id = sel_quote_hist
                            st.rerun()

                    search_name = f"HIST_{q_no}_{cust}"
                    fid, fname, pid = search_file_in_drive_by_name(search_name)
                    if fid and st.button(f"📥 Tải file chi tiết: {fname}"):
                        fh = download_from_drive(fid)
                        if fh:
                             if fname.lower().endswith('.csv'): st.dataframe(pd.read_csv(fh), use_container_width=True)
                             else: st.dataframe(pd.read_excel(fh), use_container_width=True)
        else: st.info("Chưa có lịch sử.")

    st.divider()
    st.subheader("TÍNH TOÁN & LÀM BÁO GIÁ")
    
    # 3. INPUTS CHÍNH
    c1, c2, c3 = st.columns([2, 2, 1])
    cust_db = load_data("crm_customers")
    cust_name = c1.selectbox("Chọn Khách Hàng", [""] + cust_db["short_name"].tolist() if not cust_db.empty else [])
    quote_no = c2.text_input("Số Báo Giá", key="q_no")
    
    with c3:
        if st.button("🔄 Reset", type="primary"): 
            st.session_state.quote_df = pd.DataFrame()
            st.rerun()

    with st.expander("Cấu hình chi phí (%)", expanded=True):
        cols = st.columns(7)
        keys = ["end", "buy", "tax", "vat", "pay", "mgmt", "trans"]
        params = {}
        for i, k in enumerate(keys):
            default_val = st.session_state.get(f"pct_{k}", "0")
            val = cols[i].text_input(k.upper(), value=default_val, key=f"input_{k}")
            st.session_state[f"pct_{k}"] = val
            params[k] = local_parse_money(val) 

    # 4. MATCHING & FORMULA
    cf1, cf2 = st.columns([1, 2])
    rfq = cf1.file_uploader("Upload RFQ (xlsx)", type=["xlsx"])
    
    if rfq and cf2.button("🔍 Matching (3 Biến Tuyệt Đối)"):
        st.session_state.quote_df = pd.DataFrame()
        db = load_data("crm_purchases")
        if not db.empty:
            db_recs = db.to_dict('records')
            df_rfq = pd.read_excel(rfq, dtype=str).fillna("")
            res = []
            cols_found = {clean_key(c): c for c in df_rfq.columns}
            
            for i, r in df_rfq.iterrows():
                def get_val(kws):
                    for k in kws:
                        if cols_found.get(k): return safe_str(r[cols_found.get(k)])
                    return ""
                
                code = get_val(["item code", "code", "part number"])
                name = get_val(["item name", "name", "description"])
                specs = get_val(["specs", "quy cách"])
                qty = local_parse_money(get_val(["q'ty", "qty", "quantity"])) or 1.0
                
                norm_code = normalize_match_str(code)
                norm_name = normalize_match_str(name)
                norm_specs = normalize_match_str(specs)
                
                match = None
                for cand in db_recs:
                    db_code = normalize_match_str(cand.get('item_code'))
                    db_name = normalize_match_str(cand.get('item_name'))
                    db_specs = normalize_match_str(cand.get('specs'))
                    if (db_code == norm_code) and (db_name == norm_name) and (db_specs == norm_specs):
                        match = cand; break
                
                warning = "⚠️ Không tìm thấy data"
                if match: warning = ""
                
                buying_rmb = to_float(match['buying_price_rmb']) if match else 0
                exchange_rate = to_float(match['exchange_rate']) if match else 0
                buying_vnd = to_float(match['buying_price_vnd']) if match else 0
                if buying_rmb > 0 and exchange_rate > 0: buying_vnd = buying_rmb * exchange_rate

                item = {
                    "Select": False, "No": i+1, "Cảnh báo": warning,
                    "Item code": code, "Item name": name, "Specs": specs, "Q'ty": qty,
                    "Buying price(RMB)": buying_rmb, "Exchange rate": exchange_rate, "Buying price(VND)": buying_vnd,
                    "AP price(VND)": 0, "Unit price(VND)": 0, "Total price(VND)": 0,
                    "Leadtime": match['leadtime'] if match else "", "Supplier": match['supplier_name'] if match else ""
                }
                res.append(item)
            st.session_state.quote_df = pd.DataFrame(res)
            st.rerun()

    # --- KHỐI FORMULA BUTTONS ---
    c_form1, c_form2 = st.columns(2)
    with c_form1:
        ap_f = st.text_input("Formula AP (vd: BUY*1.1)", key="f_ap")
        if st.button("Apply AP"):
            if not st.session_state.quote_df.empty:
                for idx, row in st.session_state.quote_df.iterrows():
                    buy = local_parse_money(row.get("Buying price(VND)", 0))
                    ap = local_parse_money(row.get("AP price(VND)", 0))
                    new_ap = local_eval_formula(ap_f, buy, ap)
                    st.session_state.quote_df.at[idx, "AP price(VND)"] = new_ap
                    
                    old_unit = local_parse_money(row.get("Unit price(VND)", 0))
                    markup = old_unit/ap if ap > 0 else 1.1
                    if new_ap > 0:
                          st.session_state.quote_df.at[idx, "Unit price(VND)"] = new_ap * markup
            
                st.toast("✅ Đã áp dụng công thức AP!", icon="✨")
                st.rerun()
                
    with c_form2:
        unit_f = st.text_input("Formula Unit (vd: AP*1.2)", key="f_unit")
        if st.button("Apply Unit"):
            if not st.session_state.quote_df.empty:
                for idx, row in st.session_state.quote_df.iterrows():
                    buy = local_parse_money(row.get("Buying price(VND)", 0))
                    ap = local_parse_money(row.get("AP price(VND)", 0))
                    new_unit = local_eval_formula(unit_f, buy, ap)
                    st.session_state.quote_df.at[idx, "Unit price(VND)"] = new_unit
                st.toast("✅ Đã áp dụng công thức Unit Price!", icon="✨")
                st.rerun()

    # 5. HIỂN THỊ BẢNG TAB 3
    if not st.session_state.quote_df.empty:
        st.session_state.quote_df = recalculate_quote_logic(st.session_state.quote_df, params)
        
        ordered_cols = [
            "Select", "No", "Cảnh báo", "Item code", "Item name", "Specs", "Q'ty",
            "Buying price(RMB)", "Total buying price(rmb)", "Exchange rate",
            "Buying price(VND)", "Total buying price(VND)",
            "AP price(VND)", "AP total price(VND)",
            "Unit price(VND)", "Total price(VND)", "GAP",
            "End user(%)", "Buyer(%)", "Import tax(%)", "VAT",
            "Transportation", "Management fee(%)", "Payback(%)",
            "Profit(VND)", "Profit(%)", "Supplier", "Leadtime"
        ]
        display_cols = [c for c in ordered_cols if c in st.session_state.quote_df.columns]
        df_display = st.session_state.quote_df[display_cols].copy()
        
        # --- TOTAL ROW ---
        total_row = {"Select": False, "No": "TOTAL", "Cảnh báo": "", "Item code": "", "Item name": "", "Specs": "", "Q'ty": 0}
        sum_cols = ["Q'ty", "Buying price(RMB)", "Total buying price(rmb)", 
                    "Buying price(VND)", "Total buying price(VND)",
                    "AP price(VND)", "AP total price(VND)", 
                    "Unit price(VND)", "Total price(VND)", "GAP",
                    "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", 
                    "Transportation", "Management fee(%)", "Payback(%)", "Profit(VND)"]
        for c in sum_cols:
            if c in st.session_state.quote_df.columns:
                val = st.session_state.quote_df[c].apply(local_parse_money).sum()
                if "RMB" in c or "rmb" in c: total_row[c] = local_fmt_rmb(val)
                else: total_row[c] = local_fmt_vnd(val)
        
        t_profit = local_parse_money(total_row.get("Profit(VND)", "0"))
        t_price = local_parse_money(total_row.get("Total price(VND)", "0"))
        total_row["Profit(%)"] = f"{(t_profit / t_price * 100) if t_price > 0 else 0:.1f}%"
        
        df_display = pd.concat([df_display, pd.DataFrame([total_row])], ignore_index=True)

        cols_vnd_fmt = ["Buying price(VND)", "Total buying price(VND)", "AP price(VND)", "AP total price(VND)", 
                        "Unit price(VND)", "Total price(VND)", "GAP", "Profit(VND)", 
                        "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", "Transportation", "Management fee(%)", "Payback(%)"]
        for c in cols_vnd_fmt:
            if c in df_display.columns: 
                df_display[c] = df_display.apply(lambda x: local_fmt_vnd(x[c]) if x["No"] != "TOTAL" else x[c], axis=1)
            
        cols_rmb_fmt = ["Buying price(RMB)", "Total buying price(rmb)", "Exchange rate"]
        for c in cols_rmb_fmt:
            if c in df_display.columns: 
                df_display[c] = df_display.apply(lambda x: local_fmt_rmb(x[c]) if x["No"] != "TOTAL" else x[c], axis=1)
        
        st.markdown("---")
        col_cfg = {
            "Select": st.column_config.CheckboxColumn("✅", width="small"),
            "Cảnh báo": st.column_config.TextColumn("Cảnh báo", disabled=True, width="small"),
            "No": st.column_config.TextColumn("No", width="small"),
            "Q'ty": st.column_config.NumberColumn("Q'ty", format="%d", width="small"),
            "Profit(%)": st.column_config.TextColumn("Profit(%)", width="small"), 
            "Supplier": st.column_config.TextColumn("Supplier", width="medium"),
            "Leadtime": st.column_config.TextColumn("Leadtime", width="small"),
            "Item code": st.column_config.TextColumn("Item code", width="medium"),
            "Item name": st.column_config.TextColumn("Item name", width="medium"),
            "Specs": st.column_config.TextColumn("Specs", width="medium")
        }
        for c in cols_vnd_fmt + cols_rmb_fmt:
            col_cfg[c] = st.column_config.TextColumn(c, width="small")

        edited_df = st.data_editor(
            df_display,
            column_config=col_cfg,
            use_container_width=True, height=600, key="main_editor", hide_index=True
        )

        total_q_val_main = st.session_state.quote_df["Total price(VND)"].apply(local_parse_money).sum()
        st.markdown(f'<div style="text-align: right; font-size: 20px; font-weight: bold; margin-top: 10px; padding: 10px; background-color: #f0f2f6; border-radius: 5px; color: #31333F;">💰 TỔNG CỘNG: {local_fmt_vnd(total_q_val_main)} VND</div>', unsafe_allow_html=True)
        
        # Sync Logic
        df_new_data = edited_df[edited_df["No"] != "TOTAL"].reset_index(drop=True)
        if not df_new_data.empty and len(df_new_data) == len(st.session_state.quote_df):
            data_changed = False
            for i, row_new in df_new_data.iterrows():
                row_old = st.session_state.quote_df.iloc[i]
                
                if "AP price(VND)" in row_new:
                    new_ap = local_parse_money(row_new["AP price(VND)"])
                    old_ap = local_parse_money(row_old.get("AP price(VND)", 0))
                    if abs(new_ap - old_ap) > 1.0:
                        st.session_state.quote_df.at[i, "AP price(VND)"] = new_ap
                        old_unit = local_parse_money(row_old.get("Unit price(VND)", 0))
                        markup = 1.1
                        if old_ap > 0: markup = old_unit / old_ap
                        st.session_state.quote_df.at[i, "Unit price(VND)"] = new_ap * markup
                        data_changed = True
                        continue

                check_cols = ["Q'ty", "Buying price(VND)", "Unit price(VND)", "Buying price(RMB)"]
                for col in check_cols:
                    if col in row_new:
                        new_val = local_parse_money(row_new[col])
                        old_val = local_parse_money(row_old.get(col, 0))
                        if abs(new_val - old_val) > 1.0:
                            st.session_state.quote_df.at[i, col] = new_val
                            data_changed = True
                
                if "Select" in row_new and row_new["Select"] != row_old.get("Select", False):
                     st.session_state.quote_df.at[i, "Select"] = row_new["Select"]
                     data_changed = True
                if "Item name" in row_new and str(row_new["Item name"]) != str(row_old.get("Item name","")):
                     st.session_state.quote_df.at[i, "Item name"] = str(row_new["Item name"])
                     data_changed = True

            if data_changed:
                st.session_state.quote_df = recalculate_quote_logic(st.session_state.quote_df, params)
                st.rerun()

    # 6. TOOLBAR & SAVING
    if not st.session_state.quote_df.empty:
        st.divider()
        c_rev, c_sv = st.columns([1, 1])
        with c_rev:
            st.markdown('<div class="dark-btn">', unsafe_allow_html=True)
            if st.button("🔍 REVIEW BÁO GIÁ"): st.session_state.show_review = True
            st.markdown('</div>', unsafe_allow_html=True)
        
        if st.session_state.get('show_review', False):
            st.write("### 📋 BẢNG REVIEW")
            cols_review = ["No", "Item code", "Item name", "Specs", "Q'ty", "Unit price(VND)", "Total price(VND)", "Leadtime"]
            valid_cols = [c for c in cols_review if c in st.session_state.quote_df.columns]
            df_review = st.session_state.quote_df[valid_cols].copy()
            
            total_qty = df_review["Q'ty"].apply(local_parse_money).sum() if "Q'ty" in df_review.columns else 0
            total_unit = df_review["Unit price(VND)"].apply(local_parse_money).sum() if "Unit price(VND)" in df_review.columns else 0
            total_price = df_review["Total price(VND)"].apply(local_parse_money).sum() if "Total price(VND)" in df_review.columns else 0
            
            if "Unit price(VND)" in df_review.columns: df_review["Unit price(VND)"].apply(local_fmt_vnd)
            if "Total price(VND)" in df_review.columns: df_review["Total price(VND)"].apply(local_fmt_vnd)
            
            rev_total = {"No": "TOTAL", "Q'ty": total_qty, "Unit price(VND)": local_fmt_vnd(total_unit), "Total price(VND)": local_fmt_vnd(total_price)}
            df_review = pd.concat([df_review, pd.DataFrame([rev_total])], ignore_index=True)
            
            st.dataframe(df_review, use_container_width=True, hide_index=True)
            total_q_val = st.session_state.quote_df["Total price(VND)"].apply(local_parse_money).sum()
            st.markdown(f'<div class="total-view">💰 TỔNG CỘNG: {local_fmt_vnd(total_q_val)} VND</div>', unsafe_allow_html=True)
            
            st.markdown('<div class="dark-btn">', unsafe_allow_html=True)
            if st.button("📤 XUẤT BÁO GIÁ (Excel)"):
                if not cust_name: st.error("Chưa chọn khách hàng!")
                else:
                    try:
                        df_tmpl = load_data("crm_templates")
                        match_tmpl = df_tmpl[df_tmpl['template_name'].astype(str).str.contains("AAA-QUOTATION", case=False, na=False)]
                        if match_tmpl.empty: st.error("Không tìm thấy template 'AAA-QUOTATION'!")
                        else:
                            tmpl_id = match_tmpl.iloc[0]['file_id']
                            fh = download_from_drive(tmpl_id)
                            if not fh: st.error("Lỗi tải template!")
                            else:
                                wb = load_workbook(fh); ws = wb.active
                                start_row = 11
                                first_leadtime = st.session_state.quote_df.iloc[0]['Leadtime'] if not st.session_state.quote_df.empty else ""
                                ws['H8'] = safe_str(first_leadtime)
                                for idx, row in st.session_state.quote_df.iterrows():
                                    r = start_row + idx
                                    ws[f'A{r}'] = row['No']
                                    ws[f'C{r}'] = row['Item code']
                                    ws[f'D{r}'] = row['Item name']
                                    ws[f'E{r}'] = row['Specs']
                                    ws[f'F{r}'] = local_parse_money(row["Q'ty"])
                                    ws[f'G{r}'] = local_parse_money(row["Unit price(VND)"])
                                    ws[f'H{r}'] = local_parse_money(row["Total price(VND)"])
                                out = io.BytesIO(); wb.save(out); out.seek(0)
                                curr_year = datetime.now().strftime("%Y")
                                curr_month = datetime.now().strftime("%b").upper()
                                fname = f"QUOTE_{quote_no}_{cust_name}_{int(time.time())}.xlsx"
                                path_list = ["QUOTATION_HISTORY", cust_name, curr_year, curr_month]
                                lnk, _ = upload_to_drive_structured(out, path_list, fname)
                                st.success(f"✅ Đã xuất báo giá: {fname}")
                                st.markdown(f"📂 [Mở Folder]({lnk})", unsafe_allow_html=True)
                                st.download_button("📥 Tải File", data=out, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    except Exception as e: st.error(f"Lỗi xuất Excel: {e}")
            st.markdown('</div>', unsafe_allow_html=True)

            with c_sv:
                st.markdown('<div class="dark-btn">', unsafe_allow_html=True)
                if st.button("💾 LƯU LỊCH SỬ (QUAN TRỌNG)"):
                    if cust_name:
                        # 1. Params
                        clean_params = {}
                        for k, v in params.items():
                            if isinstance(v, float) and (np.isnan(v) or np.isinf(v)): clean_params[k] = 0.0
                            else: clean_params[k] = v
                        
                        # 2. Config JSON
                        full_data_list = []
                        for r in st.session_state.quote_df.to_dict('records'):
                            clean_row = {}
                            for k_row, v_row in r.items():
                                if isinstance(v_row, (pd.Timestamp, datetime)): clean_row[k_row] = str(v_row)
                                else: clean_row[k_row] = v_row
                            full_data_list.append(clean_row)
                        
                        config_json = json.dumps({"params": clean_params, "full_data": full_data_list})
                        
                        # 3. Insert SQL
                        recs = []
                        history_id_gen = f"{cust_name}_{int(time.time())}"
                        now_str = datetime.now().strftime("%Y-%m-%d")

                        for r in st.session_state.quote_df.to_dict('records'):
                            def get_num(key): return local_parse_money(r.get(key, 0))
    
                            recs.append({
                                "history_id": history_id_gen, 
                                "date": now_str,
                                "quote_no": quote_no, 
                                "customer": cust_name,
                                "item_code": str(r.get("Item code", "")), 
                                "item_name": str(r.get("Item name", "")), 
                                "specs": str(r.get("Specs", "")),    
                                "qty": clean_number_for_db(get_num("Q'ty")),
                                "unit_price": get_num("Unit price(VND)"),
                                "total_price_vnd": get_num("Total price(VND)"),
                                "profit_vnd": get_num("Profit(VND)"),
                                
                                "buying_price_rmb": get_num("Buying price(RMB)"),
                                "total_buying_price_rmb": get_num("Total buying price(rmb)"),
                                "exchange_rate": get_num("Exchange rate"),
                                "buying_price_vnd": get_num("Buying price(VND)"),
                                "total_buying_price_vnd": get_num("Total buying price(VND)"),
                                "ap_price_vnd": get_num("AP price(VND)"),
                                "ap_total_price_vnd": get_num("AP total price(VND)"),
                                "gap": get_num("GAP"),
                                
                                "end_user_pct": clean_params.get("end", 0),
                                "buyer_pct": clean_params.get("buy", 0),
                                "import_tax_pct": clean_params.get("tax", 0),
                                "vat_pct": clean_params.get("vat", 0), 
                                "vat_money": get_num("VAT"),
                                "transportation": get_num("Transportation"),
                                "management_fee_pct": clean_params.get("mgmt", 0),
                                "payback_pct": clean_params.get("pay", 0),
                                
                                "profit_pct_display": str(r.get("Profit(%)", "")),
                                "total_cogs_vnd": get_num("Total price(VND)") - get_num("Profit(VND)"), 
                                "config_data": config_json 
                            })
                        
                        try:
                            supabase.table("crm_quotations_log").insert(recs).execute()
                            
                            df_save = st.session_state.quote_df.copy()
                            required_cols_order = [
                                "Item code", "Item name", "Specs", "Q'ty", 
                                "Buying price(RMB)", "Total buying price(rmb)", "Exchange rate", 
                                "Buying price(VND)", "Total buying price(VND)", 
                                "AP price(VND)", "AP total price(VND)", 
                                "Unit price(VND)", "Total price(VND)", 
                                "GAP", 
                                "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", 
                                "Transportation", "Management fee(%)", "Payback(%)", 
                                "Profit(VND)", "Profit(%)", "Supplier", "Leadtime"
                            ]
                            for c in required_cols_order:
                                if c not in df_save.columns: df_save[c] = ""
                            df_save = df_save[required_cols_order]

                            xlsx_buffer = io.BytesIO()
                            df_save.to_excel(xlsx_buffer, index=False)
                            xlsx_buffer.seek(0)
                            
                            xlsx_name = f"HIST_{quote_no}_{cust_name}_{int(time.time())}.xlsx"
                            curr_year = datetime.now().strftime("%Y"); curr_month = datetime.now().strftime("%b").upper()
                            path_list_hist = ["QUOTATION_HISTORY", cust_name, curr_year, curr_month]
                            lnk, _ = upload_to_drive_structured(xlsx_buffer, path_list_hist, xlsx_name)
                            
                            df_cfg = pd.DataFrame([clean_params])
                            cfg_buffer = io.BytesIO()
                            df_cfg.to_excel(cfg_buffer, index=False)
                            cfg_buffer.seek(0)
                            cfg_name = f"CONFIG_{quote_no}_{cust_name}_{int(time.time())}.xlsx"
                            upload_to_drive_structured(cfg_buffer, path_list_hist, cfg_name)
                            
                            st.success("✅ Đã lưu Lịch sử + Cơ cấu chi phí thành công!")
                            st.markdown(f"📂 [Folder Lịch Sử]({lnk})", unsafe_allow_html=True)
                        except Exception as e:
                            st.error(f"Lỗi lưu DB: {e}")
                    else: st.error("Chọn khách!")
                st.markdown('</div>', unsafe_allow_html=True)
import pandas as pd
import streamlit as st
import numpy as np
import io
from datetime import datetime
from openpyxl import Workbook
import time

# =============================================================================
# --- HELPER FUNCTIONS (XỬ LÝ FORMAT TIỀN TỆ) ---
# =============================================================================
def local_parse_money(value):
    """
    Chuyển chuỗi '1,200,000' hoặc '1,200.50' thành số float (1200000.0).
    Nếu lỗi trả về 0.0
    """
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Loại bỏ dấu phẩy, khoảng trắng
        clean_val = value.replace(',', '').strip()
        try:
            return float(clean_val)
        except ValueError:
            return 0.0
    return 0.0

def local_fmt_vnd(value):
    """Format số thành chuỗi integer có dấu phẩy: 1,200,000"""
    try:
        val = float(value)
        return "{:,.0f}".format(val)
    except:
        return "0"

def local_fmt_rmb(value):
    """Format số thành chuỗi float 2 số lẻ: 1,200.50"""
    try:
        val = float(value)
        return "{:,.2f}".format(val)
    except:
        return "0.00"
# =============================================================================
# --- TAB 4: QUẢN LÝ PO (FULL VERSION - ĐÃ THÊM UPLOAD PDF/IMG CHO KHÁCH) ---
# =============================================================================
# LƯU Ý: Đảm bảo không có lệnh 'try:' nào bị bỏ dở ngay phía trên dòng này
with t4:
    # --- 1. CÁC HÀM HỖ TRỢ (HELPER FUNCTIONS) ---
    def local_parse_money(val):
        """Chuyển chuỗi có dấu phẩy/text về số Float để tính toán"""
        try:
            if pd.isna(val) or str(val).strip() == "": return 0.0
            if isinstance(val, (int, float)): return float(val)
            s = str(val).replace(",", "").replace("%", "").strip()
            return float(s)
        except: return 0.0

    def local_fmt_vnd(val):
        """Format số thành chuỗi 1,200,000"""
        try:
            if pd.isna(val) or val == "": return "0"
            return "{:,.0f}".format(round(float(val)))
        except: return str(val)

    def local_fmt_rmb(val):
        """Format số thành chuỗi 1,200.00"""
        try:
            if pd.isna(val) or val == "": return "0.00"
            return "{:,.2f}".format(float(val))
        except: return str(val)
    
    def to_float(val): return local_parse_money(val)

    def normalize_match_str(val):
        if pd.isna(val) or val is None: return ""
        s = str(val).lower().strip()
        s = re.sub(r'[^a-z0-9]', '', s) 
        return s
    
    def get_history_config(record):
        try:
            if record.get('config_data'):
                cfg = json.loads(record['config_data'])
                return cfg.get('params', {}) 
        except: pass
        return {}

    def get_deep_history_info(record, target_code):
        supp, lead = "", ""
        try:
            if record.get('supplier_name'): supp = str(record.get('supplier_name'))
            if record.get('leadtime'): lead = str(record.get('leadtime'))
            
            if (not supp or not lead) and record.get('config_data'):
                cfg = json.loads(record['config_data'])
                full_data = cfg.get('full_data', [])
                if full_data:
                    norm_target = normalize_match_str(target_code)
                    for item in full_data:
                        if normalize_match_str(item.get('Item code', '')) == norm_target:
                            if not supp: supp = str(item.get('Supplier', ''))
                            if not lead: lead = str(item.get('Leadtime', ''))
                            break
        except: pass
        return supp, lead

    # --- 2. LOGIC TÍNH TOÁN (CORE) ---
    def recalculate_po_logic_final(df):
        if df.empty: return df
        for idx, row in df.iterrows():
            try:
                if str(row.get("No")) == "TOTAL": continue
                qty = local_parse_money(row.get("Q'ty", 0))
                
                # Giá Mua
                buy_vnd = local_parse_money(row.get("Buying price(VND)", 0))
                buy_rmb = local_parse_money(row.get("Buying price(RMB)", 0))
                ex_rate = local_parse_money(row.get("Exchange rate", 0))
                if buy_vnd == 0 and buy_rmb > 0 and ex_rate > 0:
                    buy_vnd = round(buy_rmb * ex_rate, 0)

                total_buy_vnd = round(buy_vnd * qty, 0)
                total_buy_rmb = round(buy_rmb * qty, 2)

                # Giá Bán & AP
                ap_vnd = local_parse_money(row.get("AP price(VND)", 0))
                ap_total = round(ap_vnd * qty, 0)

                unit_price = local_parse_money(row.get("Unit price(VND)", 0))
                total_sell = round(unit_price * qty, 0)
                gap = total_sell - ap_total

                # Chi phí
                val_imp_tax = local_parse_money(row.get("Import tax(%)", 0))
                val_end = local_parse_money(row.get("End user(%)", 0))
                val_buyer = local_parse_money(row.get("Buyer(%)", 0))
                val_vat = local_parse_money(row.get("VAT", 0))
                val_mgmt = local_parse_money(row.get("Management fee(%)", 0))
                val_trans = local_parse_money(row.get("Transportation", 0))
                val_payback = local_parse_money(row.get("Payback(%)", 0))
                if gap <= 0: val_payback = 0.0 

                # Profit
                sum_deductions = (total_buy_vnd + gap + val_end + val_buyer + val_imp_tax + val_vat + val_trans + val_mgmt)
                val_profit = round(total_sell - sum_deductions + val_payback, 0)
                pct_profit = 0.0
                if total_sell != 0: pct_profit = (val_profit / total_sell) * 100

                # Update DF
                df.at[idx, "Buying price(VND)"] = buy_vnd
                df.at[idx, "Total buying price(rmb)"] = total_buy_rmb
                df.at[idx, "Total buying price(VND)"] = total_buy_vnd
                df.at[idx, "AP total price(VND)"] = ap_total
                df.at[idx, "Total price(VND)"] = total_sell
                df.at[idx, "GAP"] = gap
                df.at[idx, "Payback(%)"] = val_payback 
                df.at[idx, "Profit(VND)"] = val_profit
                df.at[idx, "Profit(%)"] = f"{pct_profit:.2f}%"
            except Exception: continue      
        return df

    # --- 3. GIAO DIỆN CHÍNH ---
    c_title, c_tools = st.columns([3, 2])
    with c_title:
        st.markdown("### 🔎 QUẢN LÝ PO")
    with c_tools:
        st.markdown('<div style="text-align: right;">', unsafe_allow_html=True)
        c_t1, c_t2 = st.columns([1, 1])
        with c_t1:
            if st.button("🔄 Reset", key="btn_reset_po_t4", use_container_width=True):
                st.session_state.po_main_df = pd.DataFrame()
                st.rerun()
        with c_t2:
            if st.button("🗑️ Xóa dòng", key="btn_del_rows", use_container_width=True, type="primary"):
                if 'po_main_df' in st.session_state and not st.session_state.po_main_df.empty:
                    if "✅" in st.session_state.po_main_df.columns:
                        st.session_state.po_main_df = st.session_state.po_main_df[
                            st.session_state.po_main_df["✅"] == False
                        ].reset_index(drop=True)
                        st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    if 'po_main_df' not in st.session_state: st.session_state.po_main_df = pd.DataFrame()

    # --- 3.1 INPUTS & CHỌN KHÁCH HÀNG ---
    c_in1, c_in2, c_in3 = st.columns([1, 1, 2])
    po_no_input = c_in1.text_input("Số PO", key="po_no_input_val")
    if po_no_input: st.session_state["po_no_input"] = po_no_input 
    
    cust_db = load_data("crm_customers")
    cust_name = c_in2.selectbox("Khách Hàng", [""] + cust_db["short_name"].tolist() if not cust_db.empty else [])
    
    uploaded_files = c_in3.file_uploader("Upload PO (Excel, CSV, PDF, Img)", type=["xlsx", "xls", "csv", "pdf", "png", "jpg"], accept_multiple_files=True)

    # --- 3.2 LOGIC TỰ ĐỘNG LOAD CẤU HÌNH TỪ LỊCH SỬ ---
    d_tax, d_end, d_buy, d_vat, d_mgmt, d_pay = 0.0, 0.0, 0.0, 8.0, 0.0, 0.0
    if cust_name:
        try:
            df_hist_c = load_data("crm_quotations_log")
            if not df_hist_c.empty:
                df_cust_hist = df_hist_c[df_hist_c['customer'].astype(str).str.lower() == str(cust_name).lower()]
                if not df_cust_hist.empty:
                    if 'history_id' in df_cust_hist.columns:
                        df_cust_hist = df_cust_hist.sort_values(by=["date", "history_id"], ascending=[False, False])
                    else:
                        df_cust_hist = df_cust_hist.sort_values(by="date", ascending=False)
                    
                    for _, last_rec in df_cust_hist.iterrows():
                        cfg_json = last_rec.get('config_data', '{}')
                        if cfg_json and cfg_json != "{}":
                            parsed_cfg = json.loads(cfg_json).get('params', {})
                            if parsed_cfg:
                                d_tax = float(parsed_cfg.get('tax', 0))
                                d_end = float(parsed_cfg.get('end', 0))
                                d_buy = float(parsed_cfg.get('buy', 0))
                                d_vat = float(parsed_cfg.get('vat', 8))
                                d_mgmt = float(parsed_cfg.get('mgmt', 0))
                                d_pay = float(parsed_cfg.get('pay', 0))
                                break
        except Exception: pass

    # --- 3.3 HIỂN THỊ GLOBAL CONFIG ---
    with st.expander(f"⚙️ Cấu hình Chi phí Global (Đã load auto cho: {cust_name if cust_name else 'Mặc định'})", expanded=True):
        cg1, cg2, cg3, cg4, cg5, cg6 = st.columns(6)
        k_suffix = f"_{cust_name}" if cust_name else "_def"
        g_tax = cg1.number_input("Thuế NK (%)", value=d_tax, step=1.0, key=f"gtax{k_suffix}")
        g_end = cg2.number_input("End User (%)", value=d_end, step=1.0, key=f"gend{k_suffix}")
        g_buy = cg3.number_input("Buyer (%)", value=d_buy, step=1.0, key=f"gbuy{k_suffix}")
        g_vat = cg4.number_input("VAT (%)", value=d_vat, step=1.0, key=f"gvat{k_suffix}") 
        g_mgmt = cg5.number_input("Mgmt Fee (%)", value=d_mgmt, step=1.0, key=f"gmgmt{k_suffix}")
        g_pay = cg6.number_input("Payback (%)", value=d_pay, step=1.0, key=f"gpay{k_suffix}")

    # --- ACTION: TẢI PO ---
    if st.button("🚀 Tải PO & Load Lịch Sử", key="btn_load_po_action"):
        if uploaded_files and cust_name:
            try:
                target_file = None
                for f in uploaded_files:
                    if f.name.lower().endswith(('.xls', '.xlsx', '.csv')):
                        target_file = f
                        break
                
                df_po = pd.DataFrame()
                if target_file:
                    file_name = target_file.name.lower()
                    if file_name.endswith(('.xls', '.xlsx')): df_po = pd.read_excel(target_file, dtype=str).fillna("")
                    elif file_name.endswith('.csv'): df_po = pd.read_csv(target_file, dtype=str).fillna("")
                else:
                    st.warning("⚠️ Đã nhận file nhưng không tìm thấy file Excel/CSV dữ liệu. Vui lòng kiểm tra lại.")
                
                if not df_po.empty:
                    df_hist = load_data("crm_quotations_log")
                    hist_recs = []
                    if not df_hist.empty:
                        df_hist_filtered = df_hist[df_hist['customer'].astype(str).str.lower() == str(cust_name).lower()]
                        if 'history_id' in df_hist_filtered.columns:
                            df_hist_filtered = df_hist_filtered.sort_values(by=["date", "history_id"], ascending=[False, False])
                        else:
                            df_hist_filtered = df_hist_filtered.sort_values(by="date", ascending=False)
                        hist_recs = df_hist_filtered.to_dict('records')

                    res_po = []
                    cols_map = {clean_key(c): c for c in df_po.columns}
                    def get_val_po(r, kws):
                        for k in kws: 
                            if cols_map.get(k): return safe_str(r[cols_map.get(k)])
                        return ""

                    for i, r in df_po.iterrows():
                        p_code = get_val_po(r, ["item code", "code", "mã hàng"])
                        p_name = get_val_po(r, ["item name", "name", "tên hàng"])
                        p_specs = get_val_po(r, ["specs", "spec", "specification", "quy cách"]) 
                        p_qty = local_parse_money(get_val_po(r, ["q'ty", "qty", "số lượng"]))
                        if p_qty == 0: p_qty = 1.0
                        
                        match_hist = None
                        n_code = normalize_match_str(p_code)
                        n_name = normalize_match_str(p_name)
                        n_specs = normalize_match_str(p_specs)
                        
                        for h in hist_recs:
                            h_code = normalize_match_str(h.get('item_code'))
                            h_name = normalize_match_str(h.get('item_name'))
                            h_specs = normalize_match_str(h.get('specs') or h.get('specification'))
                            if h_code == n_code and h_name == n_name and h_specs == n_specs:
                                if to_float(h.get('buying_price_vnd', 0)) > 0 or to_float(h.get('buying_price_rmb', 0)) > 0:
                                    match_hist = h; break 

                        # Init Values
                        buy_rmb=0; ex_rate=0; buy_vnd=0; ap_vnd=0; unit_price=0
                        m_tax=0; m_end=0; m_buy=0; m_vat=0; m_mgmt=0; m_trans=0; m_pay=0
                        supplier=""; leadtime=""; warning="⚠️ New Item"; hidden_cfg = {}

                        if match_hist:
                            warning = ""
                            supplier, leadtime = get_deep_history_info(match_hist, match_hist.get('item_code', ''))
                            buy_rmb = to_float(match_hist.get('buying_price_rmb', 0))
                            ex_rate = to_float(match_hist.get('exchange_rate', 0))
                            buy_vnd = to_float(match_hist.get('buying_price_vnd', 0))
                            if buy_vnd == 0 and buy_rmb > 0 and ex_rate > 0: buy_vnd = buy_rmb * ex_rate
                            
                            ap_vnd = to_float(match_hist.get('ap_price_vnd', 0))
                            unit_price = to_float(match_hist.get('unit_price', 0))
                            
                            params_hist = get_history_config(match_hist)
                            hidden_cfg = params_hist
                            
                            # Nếu Item History có config thì lấy, không thì lấy Global
                            p_tax = to_float(params_hist.get('tax', g_tax)) / 100.0
                            p_end = to_float(params_hist.get('end', g_end)) / 100.0
                            p_buy = to_float(params_hist.get('buy', g_buy)) / 100.0
                            p_vat = to_float(params_hist.get('vat', g_vat)) / 100.0 
                            p_mgmt = to_float(params_hist.get('mgmt', g_mgmt)) / 100.0
                            p_pay = to_float(params_hist.get('pay', g_pay)) / 100.0
                            
                            curr_buy_total = buy_vnd * p_qty
                            curr_ap_total = ap_vnd * p_qty
                            curr_sell_total = unit_price * p_qty
                            curr_gap = curr_sell_total - curr_ap_total
                            
                            h_old_qty = to_float(match_hist.get('qty', 1))
                            if h_old_qty <= 0: h_old_qty = 1

                            def get_hist_abs_val(keys, default_pct_val):
                                found_val = None
                                for k in keys:
                                    if k in match_hist and match_hist[k] not in [None, ""]:
                                        found_val = to_float(match_hist[k])
                                        break
                                if found_val is not None:
                                    return round((found_val / h_old_qty) * p_qty, 0)
                                return default_pct_val

                            m_trans = get_hist_abs_val(['transportation', 'trans_cost'], 0)
                            m_end = get_hist_abs_val(['end_user', 'end_user_val'], round(curr_ap_total * p_end, 0))
                            m_buy = get_hist_abs_val(['buyer', 'buyer_val'], round(curr_sell_total * p_buy, 0))
                            m_mgmt = get_hist_abs_val(['management_fee', 'mgmt_val'], round(curr_sell_total * p_mgmt, 0))
                            m_tax = get_hist_abs_val(['import_tax', 'tax_val'], round(curr_buy_total * p_tax, 0))
                            m_vat = get_hist_abs_val(['vat', 'vat_val'], round(curr_sell_total * p_vat, 0))
                            if curr_gap > 0: m_pay = round(curr_gap * p_pay, 0)
                        
                        else:
                            hidden_cfg = {'tax': g_tax, 'end': g_end, 'buy': g_buy, 'vat': g_vat, 'mgmt': g_mgmt, 'pay': g_pay}

                        row_data = {
                            "✅": False, "No": i+1, "Cảnh báo": warning,
                            "Item code": p_code, "Item name": p_name, "SPECS": p_specs,
                            "Q'ty": float(p_qty),
                            "Buying price(RMB)": float(buy_rmb), "Exchange rate": float(ex_rate),
                            "Buying price(VND)": float(buy_vnd),
                            "AP price(VND)": float(ap_vnd), "Unit price(VND)": float(unit_price),
                            "Total buying price(rmb)": 0.0, "Total buying price(VND)": 0.0,
                            "AP total price(VND)": 0.0, "Total price(VND)": 0.0, "GAP": 0.0,
                            "Import tax(%)": float(m_tax), "End user(%)": float(m_end), "Buyer(%)": float(m_buy),
                            "VAT": float(m_vat), "Management fee(%)": float(m_mgmt), "Transportation": float(m_trans),
                            "Payback(%)": float(m_pay), "Profit(VND)": 0.0, "Profit(%)": "",
                            "Supplier": supplier, "Leadtime": leadtime,
                            "_hidden_cfg": json.dumps(hidden_cfg) 
                        }
                        res_po.append(row_data)
                    
                    st.session_state.po_main_df = pd.DataFrame(res_po)
                    st.toast("✅ Đã load PO thành công!", icon="🔥")
                    st.rerun()

            except Exception as e: st.error(f"Lỗi: {e}")

    # --- 4. HIỂN THỊ & EDIT ---
    if not st.session_state.po_main_df.empty:
        # Define Columns
        cols_show = ["✅", "No", "Cảnh báo", "Item code", "Item name", "SPECS",
                     "Q'ty", "Buying price(RMB)", "Total buying price(rmb)", 
                     "Buying price(VND)", "Total buying price(VND)",
                     "AP price(VND)", "AP total price(VND)", "Unit price(VND)", "Total price(VND)", "GAP",
                     "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", "Transportation", "Payback(%)",
                     "Profit(VND)", "Profit(%)", "Supplier", "Leadtime"]
        
        for c in cols_show: 
            if c not in st.session_state.po_main_df.columns: st.session_state.po_main_df[c] = ""

        st.session_state.po_main_df = recalculate_po_logic_final(st.session_state.po_main_df)
        df_show = st.session_state.po_main_df[cols_show].copy()

        cols_fmt_vnd = ["Q'ty", "Buying price(VND)", "Total buying price(VND)", 
                        "AP price(VND)", "AP total price(VND)", "Unit price(VND)", 
                        "Total price(VND)", "GAP", "End user(%)", "Buyer(%)", 
                        "Import tax(%)", "VAT", "Transportation", "Payback(%)", "Profit(VND)"]
        cols_fmt_rmb = ["Buying price(RMB)", "Total buying price(rmb)"]

        for c in cols_fmt_vnd:
            if c in df_show.columns: df_show[c] = df_show[c].apply(local_fmt_vnd)
        for c in cols_fmt_rmb:
             if c in df_show.columns: df_show[c] = df_show[c].apply(local_fmt_rmb)

        # Total Row
        total_row = {"No": "TOTAL", "Item code": "", "Item name": ""}
        sum_cols = ["Q'ty", "Buying price(RMB)", "Total buying price(rmb)", "Buying price(VND)", 
                    "Total buying price(VND)", "AP price(VND)", "AP total price(VND)", "Unit price(VND)", 
                    "Total price(VND)", "GAP", "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", 
                    "Transportation", "Payback(%)", "Profit(VND)"]
        
        numeric_sums = {}
        for c in sum_cols:
            if c in st.session_state.po_main_df.columns:
                val_sum = st.session_state.po_main_df[c].apply(local_parse_money).sum()
                numeric_sums[c] = val_sum

        for c in sum_cols:
             if c in cols_fmt_rmb: total_row[c] = local_fmt_rmb(numeric_sums.get(c, 0))
             else: total_row[c] = local_fmt_vnd(numeric_sums.get(c, 0))
        
        t_prof = numeric_sums.get("Profit(VND)", 0)
        t_rev = numeric_sums.get("Total price(VND)", 0)
        total_po_val = t_rev 
        total_row["Profit(%)"] = f"{(t_prof/t_rev)*100:.1f}%" if t_rev > 0 else "0%"
        
        df_show = pd.concat([df_show, pd.DataFrame([total_row])], ignore_index=True)

        col_cfg = {
            "✅": st.column_config.CheckboxColumn("✅", width="small"),
            "No": st.column_config.TextColumn("No", width="small", disabled=True),
            "Cảnh báo": st.column_config.TextColumn("Cảnh báo", width="medium", disabled=True),
            "Supplier": st.column_config.TextColumn("Supplier", width="medium"),
            "Leadtime": st.column_config.TextColumn("Leadtime", width="small"),
            "Profit(%)": st.column_config.TextColumn("Profit(%)", disabled=True),
            "End user(%)": st.column_config.TextColumn("End User (Tiền)", width="small"),
            "Buyer(%)": st.column_config.TextColumn("Buyer (Tiền)", width="small"),
            "Management fee(%)": st.column_config.TextColumn("Mgmt Fee", width="small"),
            "Transportation": st.column_config.TextColumn("Vận chuyển", width="small"),
            "Import tax(%)": st.column_config.TextColumn("Thuế NK", width="small"),
        }
        for c in cols_fmt_vnd + cols_fmt_rmb:
            if c not in col_cfg:
                if c in ["Total buying price(VND)", "AP total price(VND)", "Total price(VND)", 
                         "GAP", "Profit(VND)", "Total buying price(rmb)"]:
                      col_cfg[c] = st.column_config.TextColumn(c, width="small", disabled=True)
                else: col_cfg[c] = st.column_config.TextColumn(c, width="small")

        edited_po = st.data_editor(
            df_show, column_config=col_cfg, use_container_width=True, height=600, hide_index=True, key="po_editor_fix_string_fmt"
        )

        # Sync Logic
        df_new = edited_po[edited_po["No"] != "TOTAL"].reset_index(drop=True)
        if len(df_new) == len(st.session_state.po_main_df):
            has_change = False
            if "✅" in df_new.columns and not df_new["✅"].equals(st.session_state.po_main_df["✅"]):
                 st.session_state.po_main_df["✅"] = df_new["✅"]

            for i, row_n in df_new.iterrows():
                row_o = st.session_state.po_main_df.iloc[i]
                n_qty = local_parse_money(row_n["Q'ty"])
                n_buy = local_parse_money(row_n["Buying price(VND)"])
                n_unit = local_parse_money(row_n["Unit price(VND)"])
                n_rmb = local_parse_money(row_n["Buying price(RMB)"])
                n_ap = local_parse_money(row_n["AP price(VND)"]) 
                o_qty = local_parse_money(row_o["Q'ty"])
                o_buy = local_parse_money(row_o["Buying price(VND)"])
                o_unit = local_parse_money(row_o["Unit price(VND)"])
                o_rmb = local_parse_money(row_o["Buying price(RMB)"])
                o_ap = local_parse_money(row_o["AP price(VND)"])

                if (abs(n_qty-o_qty)>0.001) or (abs(n_buy-o_buy)>10) or (abs(n_unit-o_unit)>10) or (abs(n_rmb-o_rmb)>0.1) or (abs(n_ap-o_ap)>10):
                    st.session_state.po_main_df.at[i, "Q'ty"] = n_qty
                    st.session_state.po_main_df.at[i, "Buying price(VND)"] = n_buy
                    st.session_state.po_main_df.at[i, "Unit price(VND)"] = n_unit
                    st.session_state.po_main_df.at[i, "Buying price(RMB)"] = n_rmb
                    st.session_state.po_main_df.at[i, "AP price(VND)"] = n_ap 
                    try:
                        cfg = json.loads(row_o.get("_hidden_cfg", "{}"))
                        if not cfg: cfg = {'tax': g_tax, 'end': g_end, 'buy': g_buy, 'vat': g_vat, 'mgmt': g_mgmt, 'pay': g_pay}

                        if abs(n_qty-o_qty)>0.001: 
                            ratio = n_qty / o_qty if o_qty > 0 else 1
                            for sc in ["Transportation", "End user(%)", "Buyer(%)", "Management fee(%)"]:
                                old_val = local_parse_money(row_o.get(sc, 0))
                                st.session_state.po_main_df.at[i, sc] = round(old_val * ratio, 0)
                            
                            p_tax = to_float(cfg.get('tax', g_tax))/100.0
                            p_vat = to_float(cfg.get('vat', g_vat))/100.0
                            st.session_state.po_main_df.at[i, "Import tax(%)"] = round((n_buy * n_qty) * p_tax, 0)
                            st.session_state.po_main_df.at[i, "VAT"] = round((n_unit * n_qty) * p_vat, 0)
                    except: pass
                    has_change = True

                for k in ["Transportation", "VAT", "Import tax(%)", "End user(%)", "Buyer(%)", "Management fee(%)", "Payback(%)"]:
                    if k in row_n:
                         val_n = local_parse_money(row_n[k])
                         val_o = local_parse_money(row_o.get(k, 0))
                         if abs(val_n - val_o) > 10:
                            st.session_state.po_main_df.at[i, k] = val_n
                            has_change = True
            
            if has_change:
                st.session_state.po_main_df = recalculate_po_logic_final(st.session_state.po_main_df)
                st.rerun()

        st.markdown(f"""
        <div style="display: flex; justify-content: flex-end; margin-top: 10px;">
            <div style="padding: 10px 20px; background-color: #262730; border-radius: 5px; color: #00FF00; font-weight: bold; font-size: 20px; border: 1px solid #444;">
                💰 TỔNG CỘNG: {local_fmt_vnd(total_po_val)} VND
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # 1. REVIEW NCC
        with st.expander("📦 Review và đặt hàng nhà cung cấp (Đặt NCC)", expanded=False):
            cols_ncc = ["No", "Item code", "Item name", "SPECS", "Q'ty", 
                        "Buying price(RMB)", "Total buying price(rmb)", "Exchange rate", 
                        "Buying price(VND)", "Total buying price(VND)", "Supplier"]
            
            df_ncc_view = st.session_state.po_main_df.copy()
            if "Supplier" not in df_ncc_view.columns: df_ncc_view["Supplier"] = ""
            valid_cols = [c for c in cols_ncc if c in df_ncc_view.columns]
            df_ncc_view = df_ncc_view[valid_cols]
            
            total_row_ncc = {"No": "TOTAL", "Item code": "", "Item name": "", "SPECS": "", "Supplier": ""}
            for c in ["Q'ty", "Buying price(RMB)", "Total buying price(rmb)", "Buying price(VND)", "Total buying price(VND)"]:
                total_row_ncc[c] = df_ncc_view[c].apply(to_float).sum() if c in df_ncc_view.columns else 0.0
                
            df_ncc_fmt = df_ncc_view.copy()
            for c in ["Buying price(RMB)", "Total buying price(rmb)"]:
                if c in df_ncc_fmt.columns: df_ncc_fmt[c] = df_ncc_fmt[c].apply(local_fmt_rmb)
            for c in ["Buying price(VND)", "Total buying price(VND)"]:
                if c in df_ncc_fmt.columns: df_ncc_fmt[c] = df_ncc_fmt[c].apply(local_fmt_vnd)
            if "Q'ty" in df_ncc_fmt.columns: df_ncc_fmt["Q'ty"] = df_ncc_fmt["Q'ty"].apply(local_fmt_vnd)

            total_row_fmt = total_row_ncc.copy()
            total_row_fmt["Buying price(RMB)"] = local_fmt_rmb(total_row_ncc.get("Buying price(RMB)", 0))
            total_row_fmt["Total buying price(rmb)"] = local_fmt_rmb(total_row_ncc.get("Total buying price(rmb)", 0))
            total_row_fmt["Buying price(VND)"] = local_fmt_vnd(total_row_ncc.get("Buying price(VND)", 0))
            total_row_fmt["Total buying price(VND)"] = local_fmt_vnd(total_row_ncc.get("Total buying price(VND)", 0))
            total_row_fmt["Q'ty"] = local_fmt_vnd(total_row_ncc.get("Q'ty", 0))
            
            df_ncc_fmt = pd.concat([df_ncc_fmt, pd.DataFrame([total_row_fmt])], ignore_index=True)
            st.dataframe(df_ncc_fmt, use_container_width=True, hide_index=True)
            
            total_ncc_val = total_row_ncc.get("Total buying price(VND)", 0)
            st.markdown(f"""<div style="display: flex; justify-content: flex-end; margin-top: 10px;"><div style="padding: 10px 20px; background-color: #262730; border-radius: 5px; color: #00FF00; font-weight: bold; font-size: 20px; border: 1px solid #444;">💰 TỔNG CỘNG: {local_fmt_vnd(total_ncc_val)} VND</div></div>""", unsafe_allow_html=True)

            if st.button("🚀 Đặt hàng NCC"):
                if not st.session_state.get("po_no_input"): st.error("Thiếu số PO Khách Hàng!")
                else:
                    curr_po = st.session_state["po_no_input"]
                    grouped = st.session_state.po_main_df.groupby("Supplier")
                    curr_year = datetime.now().strftime("%Y")
                    curr_month = datetime.now().strftime("%m")
                    count_files = 0
                    for supp, group in grouped:
                        supp_name = str(supp).strip().upper() if supp else "UNKNOWN"
                        wb = Workbook(); ws = wb.active; ws.title = "PO NCC"
                        ws.append(cols_ncc)
                        group_valid = group.copy()
                        for col in cols_ncc:
                             if col not in group_valid.columns: group_valid[col] = ""
                        for r in group_valid[cols_ncc].to_dict('records'): ws.append(list(r.values()))
                        
                        sum_qty = group["Q'ty"].apply(to_float).sum()
                        sum_rmb = group["Total buying price(rmb)"].apply(to_float).sum()
                        sum_vnd = group["Total buying price(VND)"].apply(to_float).sum()
                        ws.append(["TOTAL", "", "", "", sum_qty, "", sum_rmb, "", "", sum_vnd, ""])

                        out = io.BytesIO(); wb.save(out); out.seek(0)
                        fname = f"{curr_po}-{supp_name}.xlsx"
                        path_list = ["PO_NCC", curr_year, supp_name, curr_month, str(curr_po)]
                        try: lnk, _ = upload_to_drive_structured(out, path_list, fname)
                        except: lnk = "#"
                        
                        lt_val = group.iloc[0]["Leadtime"] if "Leadtime" in group.columns else 0
                        try: eta = calc_eta(datetime.now(), lt_val)
                        except: eta = ""
                        track_rec = {"po_no": f"{curr_po}-{supp_name}", "partner": supp_name, "status": "Ordered", "order_type": "NCC", "last_update": datetime.now().strftime("%d/%m/%Y"), "eta": eta}
                        supabase.table("crm_tracking").insert([track_rec]).execute()
                        count_files += 1
                    st.success(f"✅ Đã tạo {count_files} đơn hàng NCC (Tách file) và lưu Drive!")

        # 2. REVIEW KHÁCH HÀNG
        with st.expander("👤 Review PO khách hàng và lưu PO", expanded=False):
            cols_kh = ["No", "Item code", "Item name", "SPECS", "Q'ty", "Unit price(VND)", "Total price(VND)", "Leadtime"]
            valid_cols_kh = [c for c in cols_kh if c in st.session_state.po_main_df.columns]
            df_kh_view = st.session_state.po_main_df[valid_cols_kh].copy()
            df_kh_view["Customer"] = cust_name 
            
            total_row_kh = {"No": "TOTAL", "Item code": "", "Item name": "", "SPECS": "", "Customer": "", "Leadtime": ""}
            for c in ["Q'ty", "Unit price(VND)", "Total price(VND)"]:
                total_row_kh[c] = df_kh_view[c].apply(to_float).sum() if c in df_kh_view.columns else 0.0
            
            df_kh_fmt = df_kh_view.copy()
            for c in ["Unit price(VND)", "Total price(VND)"]:
                if c in df_kh_fmt.columns: df_kh_fmt[c] = df_kh_fmt[c].apply(local_fmt_vnd)
            if "Q'ty" in df_kh_fmt.columns: df_kh_fmt["Q'ty"] = df_kh_fmt["Q'ty"].apply(local_fmt_vnd)
            
            total_row_kh_fmt = total_row_kh.copy()
            total_row_kh_fmt["Unit price(VND)"] = local_fmt_vnd(total_row_kh.get("Unit price(VND)", 0))
            total_row_kh_fmt["Total price(VND)"] = local_fmt_vnd(total_row_kh.get("Total price(VND)", 0))
            total_row_kh_fmt["Q'ty"] = local_fmt_vnd(total_row_kh.get("Q'ty", 0))
            
            df_kh_fmt = pd.concat([df_kh_fmt, pd.DataFrame([total_row_kh_fmt])], ignore_index=True)
            st.dataframe(df_kh_fmt, use_container_width=True, hide_index=True)
            
            total_kh_val = total_row_kh.get("Total price(VND)", 0)
            st.markdown(f"""<div style="display: flex; justify-content: flex-end; margin-top: 10px;"><div style="padding: 10px 20px; background-color: #262730; border-radius: 5px; color: #00FF00; font-weight: bold; font-size: 20px; border: 1px solid #444;">💰 TỔNG CỘNG: {local_fmt_vnd(total_kh_val)} VND</div></div>""", unsafe_allow_html=True)

            if st.button("💾 Lưu PO Khách Hàng"):
                if not st.session_state.get("po_no_input"): st.error("Thiếu số PO!")
                else:
                    curr_po = st.session_state["po_no_input"]
                    db_recs = []
                    eta_final = ""
                    for r in st.session_state.po_main_df.to_dict('records'):
                        eta_item = calc_eta(datetime.now(), r.get("Leadtime", 0))
                        eta_final = eta_item 
                        db_recs.append({
                            "po_number": curr_po, "customer": cust_name, "order_date": datetime.now().strftime("%d/%m/%Y"),
                            "item_code": r.get("Item code", ""), "item_name": r.get("Item name", ""), "specs": r.get("SPECS", ""), 
                            "qty": to_float(r.get("Q'ty", 0)), "unit_price": to_float(r.get("Unit price(VND)", 0)),
                            "total_price": to_float(r.get("Total price(VND)", 0)), "eta": eta_item
                        })
                    supabase.table("db_customer_orders").insert(db_recs).execute()
                    
                    curr_year = datetime.now().strftime("%Y"); curr_month = datetime.now().strftime("%m")
                    path_list = ["PO_KHACH_HANG", curr_year, str(cust_name), curr_month, str(curr_po)]
                    wb = Workbook(); ws = wb.active; ws.title = "PO CUSTOMER"; ws.append(cols_kh + ["Customer"])
                    excel_data = df_kh_view.copy()
                    for c in cols_kh: 
                        if c not in excel_data.columns: excel_data[c] = ""
                    for r in excel_data[cols_kh + ["Customer"]].to_dict('records'): ws.append(list(r.values()))
                    ws.append(["TOTAL", "", "", "", 
                               df_kh_view["Q'ty"].apply(to_float).sum() if "Q'ty" in df_kh_view else 0, 
                               df_kh_view["Unit price(VND)"].apply(to_float).sum() if "Unit price(VND)" in df_kh_view else 0, 
                               df_kh_view["Total price(VND)"].apply(to_float).sum() if "Total price(VND)" in df_kh_view else 0, "", ""])
                    out = io.BytesIO(); wb.save(out); out.seek(0)
                    fname = f"{curr_po}.xlsx"
                    try: lnk, _ = upload_to_drive_structured(out, path_list, fname)
                    except: lnk = "#"
                    
                    # --- NEW: UPLOAD CÁC FILE ĐÍNH KÈM (PDF, ẢNH...) ---
                    if uploaded_files:
                        for up_file in uploaded_files:
                            if up_file.name.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg')):
                                up_file.seek(0)
                                try: upload_to_drive_structured(up_file, path_list, up_file.name)
                                except: pass

                    track_rec = {"po_no": curr_po, "partner": cust_name, "status": "Waiting", "order_type": "KH", "last_update": datetime.now().strftime("%d/%m/%Y"), "eta": eta_final}
                    supabase.table("crm_tracking").insert([track_rec]).execute()
                    st.success("✅ Đã lưu PO Khách Hàng!"); st.markdown(f"📂 [Link File Drive]({lnk})")

        # 3. REVIEW CHI PHÍ
        with st.expander("💰 Review chi phí và lưu chi phí", expanded=False):
            cols_cost = ["No", "Item code", "Item name", "SPECS", "Q'ty", 
                         "Buying price(RMB)", "Total buying price(rmb)", "Exchange rate",
                         "Buying price(VND)", "Total buying price(VND)", 
                         "GAP", "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", 
                         "Transportation", "Management fee(%)", "Profit(%)"]
            
            valid_cols_cost = [c for c in cols_cost if c in st.session_state.po_main_df.columns]
            df_cost_view = st.session_state.po_main_df[valid_cols_cost].copy()
            
            # --- TÍNH TỔNG DÒNG (TOTAL ROW) ---
            total_row_cost = {"No": "TOTAL", "Item code": "", "Item name": "", "SPECS": "", "Profit(%)": ""}
            sum_cols_cost = ["Q'ty", "Buying price(VND)", "Total buying price(VND)", "GAP", 
                             "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", 
                             "Transportation", "Management fee(%)"]
            
            # Cộng tổng từng cột
            for c in sum_cols_cost:
                total_row_cost[c] = df_cost_view[c].apply(to_float).sum() if c in df_cost_view.columns else 0.0
            
            # --- [FIX QUAN TRỌNG] TÍNH TỔNG CHI PHÍ THEO CÔNG THỨC BẠN YÊU CẦU ---
            # Công thức: Buying(Total) + GAP + End + Buyer + Tax + VAT + Trans + Mgmt
            total_cost_val = (
                total_row_cost.get("Total buying price(VND)", 0) +
                total_row_cost.get("GAP", 0) +
                total_row_cost.get("End user(%)", 0) +
                total_row_cost.get("Buyer(%)", 0) +
                total_row_cost.get("Import tax(%)", 0) +
                total_row_cost.get("VAT", 0) +
                total_row_cost.get("Transportation", 0) +
                total_row_cost.get("Management fee(%)", 0)
            )

            # Format hiển thị bảng
            df_cost_fmt = df_cost_view.copy()
            for c in sum_cols_cost:
                 if c in df_cost_fmt.columns: df_cost_fmt[c] = df_cost_fmt[c].apply(local_fmt_vnd)
            
            # Format cột RMB nếu có
            if "Buying price(RMB)" in df_cost_fmt.columns: df_cost_fmt["Buying price(RMB)"] = df_cost_fmt["Buying price(RMB)"].apply(local_fmt_rmb)
            if "Total buying price(rmb)" in df_cost_fmt.columns: df_cost_fmt["Total buying price(rmb)"] = df_cost_fmt["Total buying price(rmb)"].apply(local_fmt_rmb)

            # Thêm dòng Total vào bảng hiển thị
            total_row_cost_fmt = total_row_cost.copy()
            for c in sum_cols_cost: total_row_cost_fmt[c] = local_fmt_vnd(total_row_cost.get(c, 0))
            df_cost_fmt = pd.concat([df_cost_fmt, pd.DataFrame([total_row_cost_fmt])], ignore_index=True)
            
            st.dataframe(df_cost_fmt, use_container_width=True, hide_index=True)
            
            # --- HIỂN THỊ CON SỐ TỔNG CỘNG ĐÃ FIX ---
            st.markdown(f"""
            <div style="display: flex; justify-content: flex-end; margin-top: 10px;">
                <div style="padding: 10px 20px; background-color: #262730; border-radius: 5px; color: #FF4B4B; font-weight: bold; font-size: 20px; border: 1px solid #444;">
                    💸 TỔNG CHI PHÍ: {local_fmt_vnd(total_cost_val)} VND
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button("💾 Lưu Chi Phí (Link Dashboard)"):
                if not st.session_state.get("po_no_input"): st.error("Thiếu số PO!")
                else:
                    curr_po = st.session_state["po_no_input"]
                    curr_year = datetime.now().strftime("%Y"); curr_month = datetime.now().strftime("%m")
                    path_list = ["CHI PHI", curr_year, str(cust_name), curr_month, str(curr_po)]
                    wb = Workbook(); ws = wb.active; ws.title = "COST"; ws.append(cols_cost)
                    excel_cost_data = df_cost_view.copy()
                    for c in cols_cost: 
                        if c not in excel_cost_data.columns: excel_cost_data[c] = ""
                    for r in excel_cost_data[cols_cost].to_dict('records'): ws.append(list(r.values()))
                    
                    # Dòng Total trong Excel cũng phải đúng thứ tự
                    vals = ["TOTAL", "", "", ""]
                    vals.append(total_row_cost.get("Q'ty", 0)); vals.append(""); vals.append(""); vals.append("")
                    vals.append(total_row_cost.get("Buying price(VND)", 0)); vals.append(total_row_cost.get("Total buying price(VND)", 0))
                    vals.append(total_row_cost.get("GAP", 0)); vals.append(total_row_cost.get("End user(%)", 0))
                    vals.append(total_row_cost.get("Buyer(%)", 0)); vals.append(total_row_cost.get("Import tax(%)", 0))
                    vals.append(total_row_cost.get("VAT", 0)); vals.append(total_row_cost.get("Transportation", 0))
                    vals.append(total_row_cost.get("Management fee(%)", 0)); vals.append("")
                    ws.append(vals)
                    
                    out = io.BytesIO(); wb.save(out); out.seek(0)
                    fname = f"{curr_po}.xlsx"
                    try: lnk, _ = upload_to_drive_structured(out, path_list, fname)
                    except: lnk = "#"
                    
                    recs_hist = []
                    for r in st.session_state.po_main_df.to_dict('records'):
                         recs_hist.append({
                            "history_id": f"PO_{curr_po}_{int(time.time())}_{r.get('Item code','')}", 
                            "date": datetime.now().strftime("%Y-%m-%d"), "quote_no": curr_po, "customer": cust_name,
                            "item_code": r.get("Item code", ""), "qty": to_float(r.get("Q'ty", 0)),
                            "unit_price": to_float(r.get("Unit price(VND)", 0)),
                            "total_price_vnd": to_float(r.get("Total price(VND)", 0)),
                            "profit_vnd": to_float(r.get("Profit(VND)", 0)), "config_data": "{}" 
                        })
                    try:
                        supabase.table("crm_shared_history").insert(recs_hist).execute()
                        st.success("✅ Đã lưu Chi phí & Lợi nhuận!"); st.markdown(f"📂 [Link File Drive]({lnk})")
                    except Exception as e: st.error(f"Lỗi lưu DB History: {e}")
import re

# =============================================================================
# --- TAB 5: TRACKING & PAYMENT (PHIÊN BẢN FINAL - TỰ ĐỘNG CHUYỂN LỊCH SỬ) ---
# =============================================================================
with t5:
    t5_1, t5_2, t5_3 = st.tabs(["📦 THEO DÕI (ACTIVE)", "💸 THANH TOÁN", "📜 LỊCH SỬ"])

    # --- 1. HÀM HỖ TRỢ CỤC BỘ ---
    def sv_fetch_data(table_name):
        """Lấy dữ liệu trực tiếp từ Server"""
        try:
            res = supabase.table(table_name).select("*").execute()
            return pd.DataFrame(res.data)
        except: return pd.DataFrame()

    def sv_clean_po(val):
        return str(val).strip().upper()

    # --- 2. XỬ LÝ LOGIC (CHẠY MỖI LẦN RELOAD) ---
    df_track = sv_fetch_data("crm_tracking")
    df_pay = sv_fetch_data("crm_payments")

    # A. Xử lý Logic Thanh Toán (Phân loại Chưa xong / Đã xong)
    df_pay_active = pd.DataFrame()
    df_pay_hist = pd.DataFrame()
    paid_set = set()

    if not df_pay.empty:
        # Lọc các đơn đã hoàn tất thanh toán
        # Điều kiện: Status="Đã nhận thanh toán" VÀ Có ngày thanh toán hợp lệ
        mask_done = (df_pay["payment_status"] == "Đã nhận thanh toán") & \
                    (df_pay["payment_date"].str.len() > 5) & \
                    (~df_pay["payment_date"].str.lower().str.contains("nan", na=True))
        
        df_pay_hist = df_pay[mask_done].copy()     # Đơn đã xong -> Qua Tab Lịch sử
        df_pay_active = df_pay[~mask_done].copy()  # Đơn chưa xong -> Ở lại Tab Thanh toán

        # Tạo tập hợp PO đã trả tiền để dùng cho Logic Tracking bên dưới
        for po in df_pay_hist["po_no"]:
            val = sv_clean_po(po)
            if val: paid_set.add(val)
    
    # B. Xử lý Logic Tracking (Phân loại Active / History)
    active_rows = []
    history_rows = []

    if not df_track.empty:
        for _, row in df_track.iterrows():
            po_raw = sv_clean_po(row.get("po_no", ""))
            otype = str(row.get("order_type", ""))
            status = str(row.get("status", ""))
            proof = str(row.get("proof_image", ""))
            
            has_proof = (len(proof) > 10) and ("nan" not in proof.lower()) 
            is_paid = po_raw in paid_set

            to_history = False
            # Logic NCC: Hàng về + Có ảnh
            if otype == "NCC" and status == "Arrived" and has_proof:
                to_history = True
            # Logic KH: Đã trả tiền (Bất kể trạng thái hàng)
            elif otype == "KH" and is_paid:
                to_history = True
            
            row["TRANG_THAI_TIEN"] = "✅ ĐÃ TRẢ" if is_paid else "❌ CHƯA"

            if to_history:
                history_rows.append(row)
            else:
                active_rows.append(row)

    df_track_active = pd.DataFrame(active_rows)
    df_track_history = pd.DataFrame(history_rows)

    # ==========================================================================
    # 4. GIAO DIỆN TAB 5.1: THEO DÕI (ACTIVE)
    # ==========================================================================
    with t5_1:
        st.subheader("5.1: ĐANG THEO DÕI")
        if st.button("🔄 CẬP NHẬT DỮ LIỆU", key="btn_f5_act", type="primary"):
            st.rerun()

        if not df_track_active.empty:
            c1, c2 = st.columns([1, 2])
            with c1:
                st.markdown("#### 🛠 Xử lý đơn hàng")
                po_list = df_track_active['po_no'].unique()
                sel_po = st.selectbox("Chọn PO", po_list, key="sel_po_act")
                
                curr = df_track_active[df_track_active['po_no'] == sel_po].iloc[0]
                
                st.info(f"Tiền: {curr.get('TRANG_THAI_TIEN')}")
                if curr.get('TRANG_THAI_TIEN') == "❌ CHƯA" and curr.get('order_type') == 'KH':
                    st.caption("👉 Chưa qua Lịch Sử do chưa hoàn tất Thanh Toán.")

                ops = ["Ordered", "Shipping", "Arrived", "Delivered", "Waiting"]
                st_now = curr.get("status", "Ordered")
                idx = ops.index(st_now) if st_now in ops else 0
                new_st = st.selectbox("Trạng thái", ops, index=idx, key="sel_st_act")
                
                up_img = st.file_uploader("Upload Proof", type=["png","jpg"], key="up_img_act")
                
                if st.button("💾 LƯU TRẠNG THÁI", key="btn_sv_act"):
                    load = {"status": new_st, "last_update": datetime.now().strftime("%d/%m/%Y")}
                    if up_img:
                        lnk, _ = upload_to_drive_simple(up_img, "CRM_PROOF", f"PRF_{sv_clean_po(sel_po)}_{int(time.time())}.png")
                        load["proof_image"] = lnk
                    
                    supabase.table("crm_tracking").update(load).eq("po_no", sel_po).execute()
                    
                    # Auto tạo phiếu thanh toán nếu KH Delivered
                    if new_st == "Delivered" and curr.get("order_type") == "KH":
                         po_clean = sv_clean_po(sel_po)
                         # Chỉ tạo nếu chưa từng tồn tại trong danh sách thanh toán (cả cũ và mới)
                         chk_exist = not df_pay.empty and (po_clean in df_pay['po_no'].apply(sv_clean_po).values)
                         if not chk_exist:
                             new_p = {
                                 "po_no": sel_po, 
                                 "partner": curr.get("partner",""),
                                 "payment_status": "Đợi xuất hóa đơn",
                                 "eta_payment": (datetime.now() + timedelta(days=30)).strftime("%d/%m/%Y")
                             }
                             supabase.table("crm_payments").insert([new_p]).execute()
                    
                    st.success("Đã lưu!")
                    time.sleep(0.5)
                    st.rerun()

                st.divider()
                if st.button("🗑️ Xóa Đơn Này", key="btn_del_act"):
                    supabase.table("crm_tracking").delete().eq("po_no", sel_po).execute()
                    st.rerun()

            with c2:
                st.dataframe(df_track_active, column_config={"proof_image": st.column_config.ImageColumn("Proof")}, use_container_width=True, hide_index=True)
        else:
            st.success("🎉 Tất cả đơn hàng đã hoàn tất (Đã qua Lịch Sử).")

    # ==========================================================================
    # 5. GIAO DIỆN TAB 5.2: THANH TOÁN (CHỈ HIỆN ĐƠN CHƯA XONG)
    # ==========================================================================
    with t5_2:
        st.subheader("5.2: QUẢN LÝ THANH TOÁN (CẦN XỬ LÝ)")
        if st.button("🔄 Tải lại", key="btn_f5_pay"):
            st.rerun()
        
        # CHỈ HIỂN THỊ DF_PAY_ACTIVE (Chưa thanh toán xong)
        if not df_pay_active.empty:
            c1, c2 = st.columns([1, 2])
            with c1:
                st.markdown("#### Cập nhật TT")
                p_list = df_pay_active['po_no'].unique()
                sel_p = st.selectbox("Chọn PO", p_list, key="sel_po_pay")
                
                row_p = df_pay_active[df_pay_active['po_no'] == sel_p].iloc[0]
                
                inv = st.text_input("Hóa Đơn", value=str(row_p.get('invoice_no','') or ''), key="inp_inv_pay")
                
                ops_p = ["Đợi xuất hóa đơn", "Đợi thanh toán", "Đã nhận thanh toán"]
                st_p = str(row_p.get('payment_status',''))
                idx_p = ops_p.index(st_p) if st_p in ops_p else 0
                new_st_p = st.selectbox("Trạng thái", ops_p, index=idx_p, key="sel_st_pay")
                
                cur_d = str(row_p.get('payment_date','') or '')
                st.caption(f"Ngày cũ: {cur_d}")
                
                if st.button("💾 LƯU & CẬP NHẬT", key="btn_sv_pay"):
                    load_p = {"invoice_no": inv, "payment_status": new_st_p}
                    
                    # Nếu chọn "Đã nhận" -> Tự điền ngày -> Nó sẽ biến mất khỏi Tab này sau khi rerun
                    if new_st_p == "Đã nhận thanh toán":
                        load_p["payment_date"] = datetime.now().strftime("%d/%m/%Y")
                    else:
                        load_p["payment_date"] = ""
                        
                    supabase.table("crm_payments").update(load_p).eq("po_no", sel_p).execute()
                    st.success("Đã lưu! (Nếu 'Đã nhận', đơn sẽ chuyển sang Tab Lịch Sử)")
                    time.sleep(1.0)
                    st.rerun()
                
                if st.button("🗑️ Xóa dòng này", key="btn_del_pay"):
                    supabase.table("crm_payments").delete().eq("po_no", sel_p).execute()
                    st.rerun()

            with c2:
                st.dataframe(df_pay_active, use_container_width=True, hide_index=True)
        else:
            st.success("👏 Tuyệt vời! Không còn đơn nào nợ tiền/chưa xử lý.")

    # ==========================================================================
    # 6. GIAO DIỆN TAB 5.3: LỊCH SỬ (TOÀN BỘ ĐÃ HOÀN TẤT)
    # ==========================================================================
    with t5_3:
        st.subheader("5.3: LỊCH SỬ HOẠT ĐỘNG")
        if st.button("🔄 Tải lại Lịch Sử", key="btn_f5_hist"):
            st.rerun()
        
        st.markdown("### 📦 Lịch sử Đơn Hàng (Tracking)")
        if not df_track_history.empty:
            st.dataframe(df_track_history, use_container_width=True, hide_index=True)
            with st.expander("🗑️ Xóa dữ liệu Tracking cũ"):
                d_sel = st.selectbox("Chọn PO xóa", df_track_history['po_no'].unique(), key="sel_del_hist")
                if st.button("Xác nhận xóa Tracking", key="btn_del_hist_confirm"):
                    supabase.table("crm_tracking").delete().eq("po_no", d_sel).execute()
                    st.rerun()
        else:
            st.info("Chưa có đơn hàng tracking trong lịch sử.")
            
        st.divider()
        
        st.markdown("### 💸 Lịch sử Thanh Toán (Payments)")
        if not df_pay_hist.empty:
            st.dataframe(df_pay_hist, use_container_width=True, hide_index=True)
            with st.expander("🛠 Chỉnh sửa / Xóa Lịch sử Thanh toán"):
                st.caption("Chỉ dùng khi cần khôi phục lại trạng thái 'Chưa thanh toán' hoặc xóa vĩnh viễn.")
                p_hist_sel = st.selectbox("Chọn PO Thanh Toán", df_pay_hist['po_no'].unique(), key="sel_p_hist_edit")
                
                c_h1, c_h2 = st.columns(2)
                with c_h1:
                    if st.button("Quay lại 'Đợi thanh toán'", key="btn_revert_pay"):
                        # Xóa ngày thanh toán để nó quay lại Tab 5.2
                        supabase.table("crm_payments").update({"payment_status": "Đợi thanh toán", "payment_date": ""}).eq("po_no", p_hist_sel).execute()
                        st.rerun()
                with c_h2:
                    if st.button("Xóa vĩnh viễn Payment", key="btn_del_pay_hist"):
                        supabase.table("crm_payments").delete().eq("po_no", p_hist_sel).execute()
                        st.rerun()
        else:
            st.info("Chưa có đơn hàng nào đã hoàn tất thanh toán.")
# =============================================================================
# --- TAB 6: MASTER DATA (RESTORED ALGORITHM V6025 - SELF HEALING IMPORT) ---
with t6:
    # CẬP NHẬT: Thêm tab "IMPORT DATA"
    tc, ts, tt, ti = st.tabs(["KHÁCH HÀNG", "NHÀ CUNG CẤP", "TEMPLATE", "IMPORT DATA"])
    
    # --- CUSTOMERS (ALGORITHM: FORCE INDEX MAPPING - BẤT CHẤP TIÊU ĐỀ) ---
    with tc:
        st.markdown("### 1. QUẢN LÝ KHÁCH HÀNG")
        df_c = load_data("crm_customers", order_by="id")
        st.dataframe(df_c, use_container_width=True, hide_index=True)
        
        st.write("---")
        st.write("📥 **Import Dữ Liệu Mới (Giữ nguyên toàn bộ Data)**")
        st.caption("💡 Mẹo: Phần mềm không quan tâm Tiêu đề cột trên Excel là gì (Giám đốc, Tên, Name...). Chỉ cần file Excel sắp xếp đúng **Thứ tự cột** giống hệt bảng trên là Import thành công 100%!")
        up_c = st.file_uploader("Upload Excel Khách Hàng", type=["xlsx"], key="up_cust_master")
        
        if up_c and st.button("🚀 CẬP NHẬT KHÁCH HÀNG (Force Map)"):
            try:
                # 1. Đọc Excel
                df = pd.read_excel(up_c, dtype=str).fillna("")
                
                # 2. Lấy danh sách tên cột chuẩn xác từ Database (bỏ id, created_at)
                db_cols = ["no", "short_name", "eng_name", "vn_name", "address_1", "address_2", "contact_person", "director", "phone", "fax", "tax_code", "destination", "payment_term"]
                if not df_c.empty:
                    db_cols = [c for c in df_c.columns if c not in ['id', 'created_at']]

                # 3. Cắt bỏ các cột rác ở tít bên phải của file Excel (nếu có)
                num_expected = len(db_cols)
                df = df.iloc[:, :num_expected]
                
                # 4. ÉP TÊN CỘT EXCEL: Bắt buộc các cột trong Excel phải mang tên chuẩn của DB
                df.columns = db_cols[:len(df.columns)]
                
                data = df.to_dict('records')
                
                if data:
                    # Clear cũ
                    supabase.table("crm_customers").delete().neq("id", 0).execute()
                    
                    # Insert mới
                    chunk_size = 100
                    for k in range(0, len(data), chunk_size):
                        batch = data[k:k+chunk_size]
                        supabase.table("crm_customers").insert(batch).execute()
                        
                    st.success(f"✅ Đã import thành công {len(data)} khách hàng (Bảo toàn 100% dữ liệu)!")
                    time.sleep(1.5)
                    st.rerun()
                else:
                    st.warning("File rỗng!")
            except Exception as e:
                st.error(f"Lỗi Import: {e}")

    # --- SUPPLIERS (ALGORITHM: FORCE INDEX MAPPING - BẤT CHẤP TIÊU ĐỀ) ---
    with ts:
        st.markdown("### 2. QUẢN LÝ NHÀ CUNG CẤP")
        df_s = load_data("crm_suppliers", order_by="id")
        st.dataframe(df_s, use_container_width=True, hide_index=True)
        
        st.write("---")
        st.write("📥 **Import Dữ Liệu Mới (Giữ nguyên toàn bộ Data)**")
        up_s = st.file_uploader("Upload Excel Nhà Cung Cấp", type=["xlsx"], key="up_supp_master")
        
        if up_s and st.button("🚀 CẬP NHẬT NHÀ CUNG CẤP (Force Map)"):
            try:
                df = pd.read_excel(up_s, dtype=str).fillna("")
                
                # Lấy danh sách cột chuẩn từ DB (bỏ id, created_at)
                db_cols_s = ["no", "short_name", "eng_name", "vn_name", "address_1", "address_2", "contact_person", "director", "phone", "fax", "tax_code", "destination", "payment_term"]
                if not df_s.empty:
                    db_cols_s = [c for c in df_s.columns if c not in ['id', 'created_at']]

                # Cắt cột thừa và Ép tên cột giống Khách hàng
                num_expected = len(db_cols_s)
                df = df.iloc[:, :num_expected]
                df.columns = db_cols_s[:len(df.columns)]
                
                data = df.to_dict('records')
                
                if data:
                    supabase.table("crm_suppliers").delete().neq("id", 0).execute()
                    
                    chunk_size = 100
                    for k in range(0, len(data), chunk_size):
                        batch = data[k:k+chunk_size]
                        supabase.table("crm_suppliers").insert(batch).execute()
                        
                    st.success(f"✅ Đã import thành công {len(data)} nhà cung cấp (Bảo toàn 100% dữ liệu)!")
                    time.sleep(1.5)
                    st.rerun()
                else:
                    st.warning("File rỗng!")
            except Exception as e:
                st.error(f"Lỗi Import: {e}")
    # --- TEMPLATE ---
    with tt:
        st.write("Upload Template Excel (Quotation)")
        up_t = st.file_uploader("File Template (.xlsx)", type=["xlsx"])
        t_name = st.text_input("Tên Template (Nhập chính xác: AAA-QUOTATION)")
        if up_t and t_name and st.button("Lưu Template"):
            lnk, fid = upload_to_drive_simple(up_t, "CRM_TEMPLATES", f"TMP_{t_name}.xlsx")
            if fid: 
                supabase.table("crm_templates").insert([{"template_name": t_name, "file_id": fid, "last_updated": datetime.now().strftime("%d/%m/%Y")}]).execute()
                st.success("OK");
                st.rerun()
        st.dataframe(load_data("crm_templates"))

    # --- IMPORT DATA (UPDATED) ---
    with ti:
        st.markdown("### 4. DỮ LIỆU IMPORT (MASTER)")
        
        # Load data
        try:
            df_i = load_data("crm_import_data", order_by="id")
            
            if not df_i.empty:
                # 1. Bỏ cột created_at (Requirement)
                if "created_at" in df_i.columns:
                    df_i = df_i.drop(columns=["created_at"])
                
                # 2. Format Price USD (Requirement: $ và 2 số thập phân)
                if "import_price_usd" in df_i.columns:
                    def fmt_price(x):
                        try:
                            # Xóa ký tự lạ, chuyển sang float rồi format
                            clean_val = str(x).replace('$', '').replace(',', '').strip()
                            if clean_val == "": return ""
                            val = float(clean_val)
                            return f"${val:,.2f}"
                        except:
                            return x
                    df_i["import_price_usd"] = df_i["import_price_usd"].apply(fmt_price)

                # 3. Search Box (Requirement)
                col_search, col_dummy = st.columns([1, 2])
                with col_search:
                    search_term = st.text_input("🔎 Tìm kiếm (Tên, HS Code, Part Number...)", key="search_import_master")
                
                if search_term:
                    # Lọc dữ liệu trên các cột quan trọng
                    mask = (
                        df_i["name_in_forwarder"].astype(str).str.contains(search_term, case=False, na=False) |
                        df_i["name_in_supplier"].astype(str).str.contains(search_term, case=False, na=False) |
                        df_i["name_in_customer"].astype(str).str.contains(search_term, case=False, na=False) |
                        df_i["hscode"].astype(str).str.contains(search_term, case=False, na=False)
                    )
                    df_i = df_i[mask]

            # 4. Hiển thị bảng (Requirement: Tăng chiều cao > 20 dòng)
            st.dataframe(
                df_i, 
                use_container_width=True, 
                hide_index=True, 
                height=800  # ~25-30 dòng
            )
        except Exception as e:
            st.info(f"Chưa có dữ liệu hoặc đang tải... ({e})")

        st.write("---")
        st.write("📥 **Import Dữ Liệu IMPORT DATA (Ghi đè toàn bộ)**")
        st.caption("Yêu cầu file có 10 cột theo đúng thứ tự: No, Name Forwarder, Name Supplier, Name Customer, Qty, UoM, Price, Tax, HSCode, Info")
        up_i = st.file_uploader("Upload Excel IMPORT DATA", type=["xlsx"], key="up_import_data_master")

        if up_i and st.button("🚀 CẬP NHẬT IMPORT DATA"):
            try:
                # 1. Read Excel
                df = pd.read_excel(up_i, dtype=str).fillna("")
                
                # 2. Mapping Columns (Force mapping by index to match DB schema exactly)
                target_cols = [
                    "no", "name_in_forwarder", "name_in_supplier", "name_in_customer", 
                    "qty", "uom", "import_price_usd", "import_tax_percent", 
                    "hscode", "clearance_custom_info"
                ]
                
                if len(df.columns) < 10:
                    st.error("File Excel không đủ 10 cột dữ liệu yêu cầu.")
                else:
                    # Lấy 10 cột đầu tiên và gán tên chuẩn DB
                    df = df.iloc[:, :10]
                    df.columns = target_cols
                    
                    data = df.to_dict('records')
                    
                    if data:
                        # 3. Clear Old Data
                        supabase.table("crm_import_data").delete().neq("id", 0).execute()
                        
                        # 4. Insert New Data (Chunking)
                        chunk_size = 100
                        for k in range(0, len(data), chunk_size):
                            batch = data[k:k+chunk_size]
                            supabase.table("crm_import_data").insert(batch).execute()
                            
                        st.success(f"✅ Đã import thành công {len(data)} dòng dữ liệu!")
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.warning("File rỗng!")
            except Exception as e:
                st.error(f"Lỗi Import: {e}")

# =============================================================================
# --- TAB 7: PROJECT MANAGEMENT (FULL VERSION - NO SHORTCUTS - SYNCED & INTERNAL TELEGRAM) ---
# =============================================================================
with t7:
    # --- 0. KHỞI TẠO BIẾN BẢO MẬT VÀ QUẢN LÝ PHIÊN ---
    # Duy trì trạng thái đăng nhập để bảo vệ các dữ liệu tài chính nhạy cảm
    if 'is_admin' not in st.session_state:
        st.session_state.is_admin = False

    # --- CẤU HÌNH TELEGRAM NỘI BỘ (CHỈ ĐIỀN TRONG CODE - AN TOÀN TUYỆT ĐỐI) ---
    # Alex điền trực tiếp Token và ID của nhóm Dự án vào đây để phần mềm tự động sử dụng
    PRJ_INTERNAL_TOKEN = "7785342410:AAHcdXRCu6qZs-M4mGowF-65AAGzc1kdXjw" 
    PRJ_INTERNAL_CHAT_ID = "-1003338498683"
    # --- 1. TẢI DỮ LIỆU TỪ CƠ SỞ DỮ LIỆU SUPABASE ---
    # Load bảng dự án, chi phí và khách hàng. Sắp xếp dự án mới nhất lên hàng đầu.
    df_projects = load_data("crm_projects", order_by="created_at", ascending=False)
    df_costs_master = load_data("crm_project_costs")
    cust_db = load_data("crm_customers")

    # Đảm bảo cột project_docs luôn tồn tại trong DataFrame để tránh lỗi hiển thị khi Database chưa có dữ liệu
    if not df_projects.empty and 'project_docs' not in df_projects.columns:
        df_projects['project_docs'] = ""

    # --- 2. GIAO DIỆN TIÊU ĐỀ VÀ PHẦN XÁC THỰC QUYỀN ADMIN ---
    c_tab_head1, c_tab_head2 = st.columns([7, 3])
    with c_tab_head1:
        st.markdown("### 🚀 TRUNG TÂM QUẢN LÝ DỰ ÁN (PROJECT COMMAND CENTER)")
    with c_tab_head2:
        # Sử dụng popover để tạo khu vực đăng nhập Admin gọn gàng, bảo mật
        with st.popover("🔑 ĐĂNG NHẬP QUẢN TRỊ VIÊN", use_container_width=True):
            if not st.session_state.is_admin:
                # Mật khẩu mặc định hệ thống Alex yêu cầu là admin123
                pwd_v7 = st.text_input("Nhập mã bảo mật hệ thống", type="password", key="pwd_tab7_v18_full_final")
                if pwd_v7 == "admin123":
                    st.session_state.is_admin = True
                    st.success("Xác thực quyền Admin thành công!")
                    time.sleep(0.5)
                    st.rerun()
            else:
                st.info("🔓 Bạn đang truy cập dưới quyền Quản trị viên")
                if st.button("🔴 ĐĂNG XUẤT KHỎI ADMIN", use_container_width=True, key="btn_logout_v18_full"):
                    st.session_state.is_admin = False
                    st.rerun()

    # --- 3. BỨC TRANH TOÀN CẢNH (KPI DASHBOARD TỔNG QUAN) ---
    if not df_projects.empty:
        df_dash_calc = df_projects.copy()
        
        # Thuật toán tính toán chi phí thực tế: Merge dữ liệu từ bảng project_costs dựa trên mã dự án
        if not df_costs_master.empty:
            df_costs_master['amount_vnd'] = pd.to_numeric(df_costs_master['amount_vnd'], errors='coerce').fillna(0)
            cost_sum = df_costs_master.groupby('project_code')['amount_vnd'].sum().reset_index(name='total_cost')
            df_dash_calc = pd.merge(df_dash_calc, cost_sum, on='project_code', how='left')
        else:
            df_dash_calc['total_cost'] = 0.0

        # Điền giá trị 0 cho những dự án chưa phát sinh chi phí
        df_dash_calc['total_cost'] = df_dash_calc['total_cost'].fillna(0)
        
        # Tính toán ngân sách thực, lợi nhuận thực và phần trăm biên lợi nhuận
        df_dash_calc['budget_val'] = df_dash_calc['budget_vnd'].apply(to_float)
        df_dash_calc['profit'] = df_dash_calc['budget_val'] - df_dash_calc['total_cost']
        df_dash_calc['profit_pct_raw'] = (df_dash_calc['profit'] / df_dash_calc['budget_val'] * 100).fillna(0)

        # Hiển thị bộ 3 thẻ Dashboard (Doanh thu, Chi phí, Lợi nhuận)
        m1, m2, m3 = st.columns(3)
        
        # LOGIC BẢO MẬT DỮ LIỆU: Chỉ Admin mới thấy các con số tài chính cụ thể
        if st.session_state.is_admin:
            m1.markdown(f"<div class='card-3d bg-sales'><h3>TỔNG DOANH THU ĐẦU TƯ</h3><h1>{fmt_num(df_dash_calc['budget_val'].sum())}</h1></div>", unsafe_allow_html=True)
            m2.markdown(f"<div class='card-3d bg-cost'><h3>TỔNG CHI PHÍ THỰC TẾ</h3><h1>{fmt_num(df_dash_calc['total_cost'].sum())}</h1></div>", unsafe_allow_html=True)
            m3.markdown(f"<div class='card-3d bg-profit'><h3>TỔNG LỢI NHUẬN DỰ KIẾN</h3><h1>{fmt_num(df_dash_calc['profit'].sum())}</h1></div>", unsafe_allow_html=True)
        else:
            m1.markdown("<div class='card-3d bg-sales'><h3>TỔNG DOANH THU ĐẦU TƯ</h3><h1>*******</h1></div>", unsafe_allow_html=True)
            m2.markdown("<div class='card-3d bg-cost'><h3>TỔNG CHI PHÍ THỰC TẾ</h3><h1>*******</h1></div>", unsafe_allow_html=True)
            m3.markdown("<div class='card-3d bg-profit'><h3>TỔNG LỢI NHUẬN DỰ KIẾN</h3><h1>*******</h1></div>", unsafe_allow_html=True)

        st.divider()

        # --- 4. BỘ LỌC TÌM KIẾM VÀ DANH SÁCH DỰ ÁN CHI TIẾT ---
        c_left, c_right = st.columns([1, 4])
        with c_left:
            st.markdown("📂 **PHÂN LOẠI THEO KHÁCH HÀNG**")
            # Tạo bộ lọc dự án theo khách hàng để thu hẹp phạm vi hiển thị
            selected_cust = st.selectbox("Chọn khách hàng:", ["TẤT CẢ"] + sorted(df_dash_calc["customer_name"].dropna().unique().tolist()), key="filter_cust_v18_full_v2")
            
            df_filtered = df_dash_calc.copy()
            if selected_cust != "TẤT CẢ":
                df_filtered = df_filtered[df_filtered["customer_name"] == selected_cust]

            st.markdown("---")
            st.markdown("🎯 **QUẢN LÝ DỰ ÁN CỤ THỂ**")
            # Dropdown để Alex chọn dự án muốn xem biểu đồ GANTT và chi phí ở phía dưới
            sel_prj_id = st.selectbox("Mã Dự Án cần xem:", df_filtered["project_code"].tolist(), key="sel_active_prj_v18_full_v2")

        with c_right:
            col_t1, col_t2 = st.columns([4, 1.5])
            col_t1.markdown("📋 **DANH SÁCH DỰ ÁN ĐANG TRIỂN KHAI TRÊN HỆ THỐNG**")

            with col_t2:
                # --- CHỨC NĂNG TẠO DỰ ÁN MỚI ---
                with st.popover("➕ TẠO DỰ ÁN MỚI", use_container_width=True):
                    st.markdown("**Điền thông tin dự án khởi tạo**")
                    p_code_n = st.text_input("Mã Dự Án (VD: HS-001-2024)", key="n_code_v18_full")
                    p_name_n = st.text_input("Tên Dự Án Chi Tiết", key="n_name_v18_full")
                    p_cust_n = st.selectbox("Khách Hàng Trực Thuộc", [""] + cust_db["short_name"].tolist() if not cust_db.empty else [], key="n_cust_v18_full")
                    p_bud_n = st.number_input("Tổng ngân sách dự kiến (VND)", min_value=0.0, step=1000000.0, key="n_bud_v18_full")
                    c_d1, c_d2 = st.columns(2)
                    p_start_n = c_d1.date_input("Ngày Bắt Đầu", value=datetime.now(), key="n_start_v18_full")
                    p_end_n = c_d2.date_input("Ngày Kết Thúc", value=datetime.now(), key="n_end_v18_full")
                    p_img_n = st.file_uploader("🖼️ Hình ảnh đại diện dự án", type=["png", "jpg", "jpeg"], key="n_img_v18_full")
                    
                    # Ô upload tài liệu đa định dạng Needing Docs khi tạo mới dự án
                    p_docs_n = st.file_uploader("📂 Hồ sơ kỹ thuật / Tài liệu kèm theo", accept_multiple_files=True, key="n_docs_v18_full")

                    if st.button("💾 LƯU THÔNG TIN KHỞI TẠO", use_container_width=True, type="primary", key="btn_save_v18_full"):
                        if p_code_n and p_name_n:
                            p_code_clean = p_code_n.strip().upper()
                            # Kiểm tra trùng mã dự án trước khi lưu vào Supabase
                            existing = supabase.table("crm_projects").select("project_code").eq("project_code", p_code_clean).execute()
                            if existing.data:
                                st.error(f"Lỗi: Mã dự án **{p_code_clean}** đã tồn tại trong hệ thống!")
                            else:
                                img_url_init = ""; doc_link_init = ""
                                if p_img_n:
                                    with st.spinner("Đang xử lý hình ảnh..."):
                                        timestamp = int(time.time())
                                        filename = f"PRJ_IMAGE_{p_code_clean}_{timestamp}.png"
                                        img_url_init, _ = upload_to_drive_simple(p_img_n, "CRM_PROJECT_IMAGES", filename)
                                
                                if p_docs_n:
                                    try:
                                        srv = get_drive_service()
                                        if srv:
                                            # Tạo cấu trúc thư mục phân cấp trên Drive: CRM_PROJECT_DOCS > Mã Dự Án
                                            path_list = ["CRM_PROJECT_DOCS", p_code_clean]
                                            folder_id = get_or_create_folder_hierarchy(srv, path_list, ROOT_FOLDER_ID)
                                            doc_link_init = f"https://drive.google.com/drive/folders/{folder_id}"
                                            for f in p_docs_n:
                                                # Upload từng file vào thư mục vừa tạo
                                                upload_to_drive_structured(f, path_list, f.name)
                                            st.success("✅ Toàn bộ tài liệu đã được tải lên Google Drive thành công!")
                                    except Exception as e_docs:
                                        st.error(f"Lỗi khi upload hồ sơ kỹ thuật: {e_docs}")

                                new_rec = {
                                    "project_code": p_code_clean,
                                    "project_name": p_name_n.strip(),
                                    "customer_name": p_cust_n.strip() if p_cust_n else None,
                                    "budget_vnd": float(p_bud_n),
                                    "start_date": str(p_start_n),
                                    "end_date": str(p_end_n),
                                    "project_image": img_url_init,
                                    "project_docs": doc_link_init, # Đồng bộ với cột project_docs mới
                                    "status": "In Progress"
                                }
                                try:
                                    supabase.table("crm_projects").insert([new_rec]).execute()
                                    st.cache_data.clear()
                                    st.success("✅ Dự án mới đã được khởi tạo thành công!"); time.sleep(0.5); st.rerun()
                                except Exception as e:
                                    st.error(f"Lỗi ghi dữ liệu Database: {e}")
                        else:
                            st.error("Vui lòng nhập đầy đủ Mã và Tên dự án!")

            # --- CHUẨN BỊ DỮ LIỆU HIỂN THỊ BẢNG DANH SÁCH DỰ ÁN ---
            df_table = df_filtered.copy()
            df_table = df_table.reset_index(drop=True)
            
            # YÊU CẦU 4: Cột số thứ tự (No) với chiều rộng nhỏ (35px) thay thế cho Checkbox
            df_table.insert(0, "No", range(1, len(df_table) + 1))

            # YÊU CẦU 1: Xử lý nội dung cột Needing Docs (Chỉ hiện Link khi đăng nhập Admin)
            def get_docs_link_v18(link):
                if link and str(link).strip() != "" and str(link).lower() != 'none':
                    return link
                return ""
            df_table['docs_render_link'] = df_table['project_docs'].apply(get_docs_link_v18)

            # Hàm Masking dữ liệu tiền tệ để bảo mật cho tài khoản thường
            def mask_data_v18_full(v, is_money=True):
                if st.session_state.is_admin:
                    return "{:,.0f}".format(float(v)) if is_money else f"{v:.1f}%"
                return "*******"

            df_table['budget_vnd_disp'] = df_table['budget_vnd'].apply(lambda x: "{:,.0f}".format(float(x)))
            df_table['total_cost_disp'] = df_table['total_cost'].apply(lambda x: mask_data_v18_full(x))
            df_table['profit_disp'] = df_table['profit'].apply(lambda x: mask_data_v18_full(x))
            df_table['% Profit'] = df_table['profit_pct_raw'].apply(lambda x: mask_data_v18_full(x, False))

            # Logic Cache Busting để cập nhật ảnh ngay lập tức khi thay đổi
            current_ts = int(time.time() * 1000)
            def make_fresh_url(url):
                if not url: return None
                if "drive.google.com" in url or "googleusercontent.com" in url:
                    separator = "&" if "?" in url else "?"
                    return f"{url}{separator}t={current_ts}"
                return url
            df_table['image_fresh'] = df_table['project_image'].apply(make_fresh_url)

            # YÊU CẦU 1: Thiết lập danh sách cột hiển thị (Ẩn tài liệu và tài chính khi chưa là Admin)
            cols_show = ['No', 'image_fresh', 'project_code', 'project_name', 'status']
            if st.session_state.is_admin:
                cols_show.append('docs_render_link') # HIỆN TÀI LIỆU KHI ĐĂNG NHẬP QUYỀN ADMIN
                cols_show.extend(['budget_vnd_disp', 'total_cost_disp', 'profit_disp', '% Profit'])

            # Render Bảng danh sách dự án bằng Data Editor
            st.data_editor(
                df_table[cols_show],
                column_config={
                    "No": st.column_config.NumberColumn("No", width=35, help="Số thứ tự"),
                    "image_fresh": st.column_config.ImageColumn("Ảnh Dự Án", width="small"),
                    "project_code": st.column_config.TextColumn("Mã DA", disabled=True),
                    "project_name": st.column_config.TextColumn("Tên Dự Án", disabled=True, width="large"),
                    "status": st.column_config.SelectboxColumn("Trạng thái", options=["In Progress", "Completed", "On Hold", "Cancelled"]),
                    "docs_render_link": st.column_config.LinkColumn("📄 Needing Docs", display_text="📂 Mở tài liệu")
                },
                use_container_width=True,
                hide_index=True,
                key="prj_table_v18_full_v3"
            )

            # --- NÚT XÓA DỰ ÁN (DÀNH CHO ADMIN) ---
            if st.session_state.is_admin:
                with st.popover("🗑️ XÓA DỰ ÁN KHỎI HỆ THỐNG", use_container_width=True):
                    st.warning("Cảnh báo: Hành động này sẽ xóa vĩnh viễn toàn bộ Nhiệm vụ và Chi phí!")
                    prj_del = st.selectbox("Chọn dự án muốn gỡ bỏ:", [""] + df_filtered["project_code"].tolist(), key="del_prj_v18_full")
                    if prj_del:
                        confirm_pass = st.text_input("Nhập 'admin' để xác nhận:", type="password", key="pwd_cfm_v18_full")
                        if st.button(f"🔥 XÁC NHẬN XÓA DỰ ÁN {prj_del}", type="primary", use_container_width=True):
                            if confirm_pass == "admin":
                                try:
                                    supabase.table("crm_projects").delete().eq("project_code", prj_del).execute()
                                    supabase.table("crm_project_tasks").delete().eq("project_code", prj_del).execute()
                                    supabase.table("crm_project_costs").delete().eq("project_code", prj_del).execute()
                                    st.cache_data.clear()
                                    st.success(f"Dự án {prj_del} đã được gỡ bỏ!"); time.sleep(1); st.rerun()
                                except Exception as e:
                                    st.error(f"Lỗi lệnh xóa: {e}")
                            else:
                                st.error("Mật khẩu xác nhận không chính xác!")

        # --- 5. QUẢN LÝ CHI TIẾT DỰ ÁN (TIẾN ĐỘ, CHI PHÍ, CÀI ĐẶT) ---
        if sel_prj_id:
            active_prj = df_dash_calc[df_dash_calc['project_code'] == sel_prj_id].iloc[0]
            st.divider()
            st.markdown(f"#### 🛠️ CHI TIẾT QUẢN LÝ: {active_prj['project_name']} ({sel_prj_id})")

            # Tải danh sách các công việc con (Tasks)
            tasks_all = load_data("crm_project_tasks")
            tasks_data = tasks_all[tasks_all["project_code"] == sel_prj_id] if not tasks_all.empty else pd.DataFrame()

            # Phân tách chức năng bằng các Tab phụ
            tab_names_v18 = ["⏳ TIẾN ĐỘ & GANTT"]
            if st.session_state.is_admin:
                tab_names_v18.extend(["💸 CHI PHÍ PHÁT SINH", "⚙️ CÀI ĐẶT DỰ ÁN"])

            sub_tabs = st.tabs(tab_names_v18)

            # --- TAB 5.1: BIỂU ĐỒ GANTT VÀ CẬP NHẬT NHIỆM VỤ (TÍCH HỢP TELEGRAM) ---
            with sub_tabs[0]:
                col_g1, col_g2 = st.columns([2, 3])
                
                # Thuật toán tính tiến độ trung bình (Avg) để hiển thị lên thanh Master Progress
                avg_progress_v18 = tasks_data['progress_pct'].apply(lambda x: to_float(str(x).split('%')[0])).mean() if not tasks_data.empty else 0
                
                with col_g1:
                    if not tasks_data.empty:
                        # Chuẩn bị dữ liệu vẽ biểu đồ GANTT bằng thư viện Altair chuyên nghiệp
                        df_gantt_v18 = tasks_data.copy()
                        df_gantt_v18['start_date'] = pd.to_datetime(df_gantt_v18['start_date'], errors='coerce')
                        df_gantt_v18['end_date'] = pd.to_datetime(df_gantt_v18['end_date'], errors='coerce')
                        
                        # Tạo dòng "Master Summary" đại diện cho tổng thời gian và tiến độ dự án
                        master_summary_row = pd.DataFrame([{
                            'task_name': f'⭐ TỔNG TIẾN ĐỘ ({avg_progress_v18:.0f}%)',
                            'start_date': pd.to_datetime(active_prj['start_date']),
                            'end_date': pd.to_datetime(active_prj['end_date']),
                            'status': 'Master'
                        }])
                        
                        # Cấu hình biểu đồ thanh cột ngang (GANTT) với màu sắc theo trạng thái
                        gantt_chart = alt.Chart(pd.concat([master_summary_row, df_gantt_v18])).mark_bar(cornerRadius=5, height=20).encode(
                            x=alt.X('start_date', title='Dòng thời gian thực hiện'),
                            x2='end_date',
                            y=alt.Y('task_name', sort=None, title='Hạng mục công việc'),
                            color=alt.Color('status', scale=alt.Scale(
                                domain=['Master', 'To-do', 'Doing', 'Review', 'Done'],
                                range=['#000000', '#D3D3D3', '#FFA500', '#3498DB', '#2ECC71']
                            ), title='Trạng thái hiện tại')
                        ).properties(height=450)
                        
                        st.altair_chart(gantt_chart, use_container_width=True)
                    else:
                        st.info("Dự án hiện chưa có nhiệm vụ con nào được khởi tạo.")

                with col_g2:
                    # Bảng chỉnh sửa trực tiếp thông tin nhiệm vụ, người làm và tiến độ
                    df_tasks_editor = tasks_data[["task_name", "assignee", "start_date", "end_date", "progress_pct", "status"]].copy() \
                        if not tasks_data.empty \
                        else pd.DataFrame(columns=["task_name", "assignee", "start_date", "end_date", "progress_pct", "status"])
                    
                    df_tasks_editor['start_date'] = pd.to_datetime(df_tasks_editor['start_date'], errors='coerce').dt.date
                    df_tasks_editor['end_date'] = pd.to_datetime(df_tasks_editor['end_date'], errors='coerce').dt.date
                    
                    edited_tasks_v18 = st.data_editor(
                        df_tasks_editor,
                        num_rows="dynamic",
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "progress_pct": st.column_config.SelectboxColumn(
                                "Tiến độ (%)",
                                options=["0% ⚪", "10% 🔴", "20% 🔴", "30% 🟠", "40% 🟠", "50% 🟡", "60% 🟡", "70% 🔵", "80% 🔵", "90% 🔵", "100% 🟢"]
                            ),
                            "status": st.column_config.SelectboxColumn(
                                "Trạng thái",
                                options=["To-do", "Doing", "Review", "Done"]
                            ),
                            "start_date": st.column_config.DateColumn("Ngày bắt đầu"),
                            "end_date": st.column_config.DateColumn("Ngày hoàn thành")
                        },
                        key=f"ed_v18_tasks_full_sync_{sel_prj_id}"
                    )
                    
                    # YÊU CẦU 2 & 3: CẬP NHẬT NHIỆM VỤ VÀ GỬI CẢNH BÁO TELEGRAM (API DÁN TRỰC TIẾP TRONG CODE)
                    if st.button("💾 LƯU TIẾN ĐỘ & GỬI THÔNG BÁO TELEGRAM", type="primary", use_container_width=True, key=f"btn_up_tasks_v18_full_{sel_prj_id}"):
                        with st.spinner("Đang đồng bộ dữ liệu dự án và quét công việc quá hạn..."):
                            try:
                                # Xóa toàn bộ task cũ của dự án này và ghi đè bằng bộ task mới vừa sửa
                                supabase.table("crm_project_tasks").delete().eq("project_code", sel_prj_id).execute()
                                
                                new_tasks_to_db = []
                                overdue_alert_msgs = [] # Danh sách các tin nhắn cảnh báo quá hạn
                                today_date = datetime.now().date()
                                
                                for r in edited_tasks_v18.to_dict('records'):
                                    if r.get('task_name'):
                                        new_tasks_to_db.append({
                                            "project_code": sel_prj_id,
                                            "task_name": r['task_name'],
                                            "assignee": r['assignee'],
                                            "start_date": str(r['start_date']) if r['start_date'] else None,
                                            "end_date": str(r['end_date']) if r['end_date'] else None,
                                            "progress_pct": r['progress_pct'],
                                            "status": r['status']
                                        })
                                        
                                        # YÊU CẦU 3: THUẬT TOÁN TỰ ĐỘNG QUÉT QUÁ HẠN
                                        if r['end_date'] and r['status'] != 'Done':
                                            task_due_date = pd.to_datetime(r['end_date']).date()
                                            if task_due_date < today_date:
                                                # Tạo nội dung tin nhắn cảnh báo cực kỳ chi tiết
                                                overdue_alert_msgs.append(
                                                    f"⚠️ <b>CẢNH BÁO QUÁ HẠN</b>\n"
                                                    f"📌 <b>Việc:</b> <i>{r['task_name']}</i>\n"
                                                    f"🏢 <b>Dự án:</b> {active_prj['project_name']}\n"
                                                    f"👤 <b>Phụ trách:</b> {r['assignee']}\n"
                                                    f"📅 <b>Ngày hết hạn:</b> {r['end_date']}"
                                                )

                                if new_tasks_to_db:
                                    supabase.table("crm_project_tasks").insert(new_tasks_to_db).execute()
                                
                                # YÊU CẦU 2: THUẬT TOÁN GỬI TELEGRAM (SỬ DỤNG API VÀ ID ĐÃ DÁN TRONG CODE)
                                if PRJ_INTERNAL_TOKEN and PRJ_INTERNAL_CHAT_ID:
                                    import requests
                                    telegram_api_url = f"https://api.telegram.org/bot{PRJ_INTERNAL_TOKEN}/sendMessage"
                                    
                                    # 1. Gửi báo cáo tiến độ dự án chung
                                    main_progress_msg = (
                                        f"🚀 <b>CẬP NHẬT DỰ ÁN</b>\n\n"
                                        f"📌 <b>Dự án:</b> {active_prj['project_name']}\n"
                                        f"🔢 <b>Mã dự án:</b> {sel_prj_id}\n"
                                        f"📊 <b>Tiến độ TB:</b> {avg_progress_v18:.0f}%\n"
                                        f"📍 <b>Trạng thái:</b> {active_prj['status']}\n\n"
                                        f"<i>Vui lòng truy cập CRM để xem chi tiết!</i>"
                                    )
                                    requests.post(telegram_api_url, json={"chat_id": PRJ_INTERNAL_CHAT_ID, "text": main_progress_msg, "parse_mode": "HTML"})
                                    
                                    # 2. Gửi các cảnh báo quá hạn cho từng công việc con nếu có
                                    for alert in overdue_alert_msgs:
                                        requests.post(telegram_api_url, json={"chat_id": PRJ_INTERNAL_CHAT_ID, "text": alert, "parse_mode": "HTML"})
                                    
                                    st.toast("✅ Đã cập nhật và gửi báo cáo Telegram thành công!", icon="🚀")
                                else:
                                    st.error("🛑 Lỗi: Chưa cấu hình Token hoặc Chat ID bên trong mã nguồn.")

                                st.success("✅ Dữ liệu tiến độ đã được đồng bộ hoàn toàn!")
                                time.sleep(0.8)
                                st.rerun()
                            except Exception as e:
                                st.error(f"Lỗi trong quá trình cập nhật: {str(e)}")

            # --- TAB 5.2: QUẢN LÝ CHI PHÍ PHÁT SINH (DÀNH CHO ADMIN) ---
            if st.session_state.is_admin:
                with sub_tabs[1]:
                    st.write(f"**Tổng Chi Phí Thực Tế (Phát sinh): {fmt_num(active_prj['total_cost'])} VND**")
                    # Tải bảng chi phí chi tiết của dự án
                    prj_costs_data = df_costs_master[df_costs_master["project_code"] == sel_prj_id] if not df_costs_master.empty else pd.DataFrame(columns=["cost_type", "amount_vnd", "ref_po", "description"])
                    df_costs_render = prj_costs_data[["cost_type", "amount_vnd", "ref_po", "description"]].copy()
                    # Format số tiền hiển thị cho dễ đọc
                    df_costs_render['amount_vnd'] = df_costs_render['amount_vnd'].apply(lambda x: "{:,.0f}".format(float(x)) if x != 0 and x is not None else "")

                    edited_costs_v18 = st.data_editor(
                        df_costs_render,
                        num_rows="dynamic",
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "cost_type": st.column_config.TextColumn("Loại chi phí đầu tư"),
                            "amount_vnd": st.column_config.NumberColumn("Số tiền thực chi (VND)", format="%,.0f"),
                            "ref_po": st.column_config.TextColumn("Hợp đồng / Số hóa đơn"),
                            "description": st.column_config.TextColumn("Mô tả chi tiết mục chi")
                        },
                        key=f"ed_v18_costs_full_final_{sel_prj_id}"
                    )

                    if st.button("💾 CẬP NHẬT CHI PHÍ DỰ ÁN", type="primary", use_container_width=True, key=f"btn_sv_costs_v18_{sel_prj_id}"):
                        with st.spinner("Đang hạch toán dữ liệu chi phí..."):
                            try:
                                supabase.table("crm_project_costs").delete().eq("project_code", sel_prj_id).execute()
                                final_costs_list = []
                                for row in edited_costs_v18.to_dict('records'):
                                    amount_val = to_float(row['amount_vnd'])
                                    if row.get('cost_type') and amount_val > 0:
                                        final_costs_list.append({
                                            "project_code": sel_prj_id,
                                            "cost_type": row['cost_type'],
                                            "amount_vnd": amount_val,
                                            "ref_po": row['ref_po'],
                                            "description": row['description']
                                        })
                                if final_costs_list:
                                    supabase.table("crm_project_costs").insert(final_costs_list).execute()
                                st.success("✅ Toàn bộ chi phí thực tế đã được lưu vào hệ thống!")
                                time.sleep(0.8); st.rerun()
                            except Exception as e:
                                st.error(f"Lỗi lưu chi phí: {str(e)}")

                # --- TAB 5.3: CÀI ĐẶT THÔNG TIN VÀ TÀI LIỆU DỰ ÁN (UPLOAD GHI ĐÈ) ---
                with sub_tabs[2]:
                    st.markdown("### ⚙️ CÀI ĐẶT DỰ ÁN & QUẢN LÝ HỒ SƠ KỸ THUẬT")
                    # Chức năng đặc biệt: Reset nhanh liên kết tài liệu để dọn dẹp Link Folder
                    if active_prj['project_docs']:
                        if st.button("🗑️ XÓA LIÊN KẾT TÀI LIỆU (LÀM TRỐNG Ô TÀI LIỆU)", use_container_width=True):
                            try:
                                supabase.table("crm_projects").update({"project_docs": ""}).eq("project_code", sel_prj_id).execute()
                                st.success("✅ Đã xóa liên kết thư mục trên bảng danh sách dự án!"); time.sleep(0.5); st.rerun()
                            except Exception as e: st.error(f"Lỗi Reset link: {e}")

                    with st.form(key=f"form_settings_v18_full_final_{sel_prj_id}"):
                        c_edit1, c_edit2 = st.columns(2)
                        up_name_full = c_edit1.text_input("Tên Dự Án (Thay đổi)", value=safe_str(active_prj['project_name']))
                        up_bud_full = c_edit2.number_input("Cập nhật Ngân Sách Đầu Tư (VND)", value=float(active_prj['budget_vnd']))
                        up_start_full = c_edit1.date_input("Ngày Bắt Đầu Lại", value=pd.to_datetime(active_prj['start_date']) if active_prj['start_date'] else datetime.now())
                        up_end_full = c_edit2.date_input("Ngày Kết Thúc Mới", value=pd.to_datetime(active_prj['end_date']) if active_prj['end_date'] else datetime.now())
                        up_status_full = c_edit1.selectbox("Cập nhật Trạng Thái", ["In Progress", "Completed", "On Hold", "Cancelled"], index=0)
                        up_img_full = st.file_uploader("Thay thế hình ảnh đại diện dự án mới", type=["png", "jpg", "jpeg"])

                        # Ô UPLOAD/UPDATE TÀI LIỆU (GHI ĐÈ FILE NẾU TRÙNG TÊN)
                        up_docs_full = st.file_uploader(
                            "📄 Cập nhật hồ sơ/tài liệu giải pháp (Word, Excel, PDF, Video... - trùng tên sẽ tự động ghi đè bản mới)",
                            type=None,
                            accept_multiple_files=True,
                            key=f"u_docs_v18_full_final_{sel_prj_id}"
                        )

                        if st.form_submit_button("💾 XÁC NHẬN CẬP NHẬT TOÀN BỘ THÔNG TIN", use_container_width=True, type="primary"):
                            update_payload = {
                                "project_name": up_name_full,
                                "budget_vnd": float(up_bud_full),
                                "start_date": str(up_start_full),
                                "end_date": str(up_end_full),
                                "status": up_status_full
                            }

                            # Xử lý cập nhật ảnh đại diện mới
                            if up_img_full:
                                with st.spinner("Đang thay thế hình ảnh dự án..."):
                                    ts_full = int(time.time())
                                    fn_full = f"PRJ_ID_{active_prj['project_code'].strip().upper()}_{ts_full}.png"
                                    fresh_img_url, _ = upload_to_drive_simple(up_img_full, "CRM_PROJECT_IMAGES", fn_full)
                                    update_payload["project_image"] = fresh_img_url

                            # XỬ LÝ ĐỒNG BỘ TÀI LIỆU LÊN DRIVE (CẬP NHẬT CỘT project_docs)
                            if up_docs_full and len(up_docs_full) > 0:
                                with st.spinner("Đang đồng bộ hồ sơ lên Google Drive (xử lý ghi đè)..."):
                                    try:
                                        srv_v18_full = get_drive_service()
                                        if srv_v18_full:
                                            # Trỏ đúng thư mục dự án trên Drive
                                            path_v18_full = ["CRM_PROJECT_DOCS", active_prj['project_code'].strip().upper()]
                                            folder_id_v18 = get_or_create_folder_hierarchy(srv_v18_full, path_v18_full, ROOT_FOLDER_ID)
                                            # Cập nhật link thư mục vào database
                                            folder_url_v18 = f"https://drive.google.com/drive/folders/{folder_id_v18}"
                                            for uploaded_file in up_docs_full:
                                                # Hàm upload structured mặc định hỗ trợ ghi đè bản mới nhất
                                                upload_to_drive_structured(uploaded_file, path_list=path_v18_full, file_name=uploaded_file.name)
                                            update_payload["project_docs"] = folder_url_v18
                                            st.success("📂 Toàn bộ hồ sơ kỹ thuật đã được đồng bộ bản mới nhất lên Drive!")
                                    except Exception as e_drive_full:
                                        st.error(f"Lỗi khi xử lý hồ sơ trên Drive: {e_drive_full}")

                            # Thực hiện lệnh cập nhật cuối cùng vào Supabase
                            try:
                                supabase.table("crm_projects").update(update_payload).eq("project_code", active_prj['project_code']).execute()
                                st.cache_data.clear()
                                st.success("✅ Toàn bộ thay đổi đã được áp dụng thành công!")
                                time.sleep(1.2); st.rerun()
                            except Exception as e:
                                st.error(f"Lỗi đồng bộ Database: {e}")

    else:
        st.info("Hệ thống chưa ghi nhận dự án nào. Vui lòng nhấn 'TẠO DỰ ÁN MỚI' để bắt đầu quản lý.")
# =============================================================================
# --- KẾT THÚC TAB 7 (FULL VERSION - ĐÃ FIX LỖI SUPERGROUP) ---
# =============================================================================
# --- TAB 8: QUẢN LÝ ISSUE (THEO DÕI SỰ CỐ / VẤN ĐỀ) ---
# =============================================================================
with t8:
    # -------------------------------------------------------------------------
    # 🤖 HỆ THỐNG AUTO-BOT NHẮC NHỞ TELEGRAM TRỰC TIẾP TRONG STREAMLIT
    # -------------------------------------------------------------------------
    if 'last_telegram_reminder' not in st.session_state:
        st.session_state.last_telegram_reminder = None

    now = datetime.now()
    # Xác định khung giờ quét (9h00-10h00 sáng và 16h30-17h30 chiều)
    is_9am_window = (now.hour == 9)
    is_430pm_window = (now.hour == 16 and now.minute >= 30) or (now.hour == 17 and now.minute <= 30)
    
    current_window = None
    if is_9am_window: current_window = f"{now.strftime('%Y-%m-%d')}_MORNING"
    elif is_430pm_window: current_window = f"{now.strftime('%Y-%m-%d')}_AFTERNOON"

    # Nếu đang trong khung giờ và chưa gửi cảnh báo trong buổi này
    if current_window and st.session_state.last_telegram_reminder != current_window:
        try:
            # Lấy các issue đang Mở
            res_remind = supabase.table("crm_issues").select("*").in_("status", ["Open", "In Progress"]).execute()
            if res_remind and res_remind.data:
                count_alerts = 0
                for iss in res_remind.data:
                    last_upd_str = iss.get("last_updated")
                    if last_upd_str:
                        try:
                            # Tính toán thời gian trễ
                            last_upd = datetime.fromisoformat(last_upd_str.replace("Z", "+00:00")).replace(tzinfo=None)
                            diff_hours = (now - last_upd).total_seconds() / 3600
                            
                            if diff_hours >= 24: # Nếu quá 24 giờ chưa cập nhật
                                msg = (
                                    f"🚨 <b>CẢNH BÁO: SỰ CỐ BỊ BỎ QUÊN QUÁ 24H</b>\n\n"
                                    f"📝 <b>Vấn đề:</b> {iss.get('description')}\n"
                                    f"👤 <b>Người phụ trách:</b> {iss.get('assignee')}\n"
                                    f"🏢 <b>Khách hàng:</b> {iss.get('customer_name')}\n"
                                    f"⏳ <b>Đã treo:</b> {int(diff_hours)} giờ chưa có cập nhật tiến độ\n\n"
                                    f"<i>👉 Yêu cầu PIC vào CRM xử lý ngay!</i>"
                                )
                                url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
                                requests.post(url, json={"chat_id": TELEGRAM_GROUP_ID, "text": msg, "parse_mode": "HTML"})
                                count_alerts += 1
                        except Exception: pass
                
                if count_alerts > 0:
                    st.toast(f"🤖 Hệ thống đã tự động gửi {count_alerts} nhắc nhở quá hạn 24h lên Telegram!", icon="🚨")
        except Exception: pass
        
        # Đánh dấu là đã quét xong trong buổi này để không gửi lặp
        st.session_state.last_telegram_reminder = current_window
    # -------------------------------------------------------------------------

    st.markdown("### ⚠️ TRUNG TÂM THEO DÕI SỰ CỐ (ISSUE TRACKING)")

    # 1. TẢI DỮ LIỆU
    df_issues = load_data("crm_issues", order_by="created_at", ascending=False)
    cust_db = load_data("crm_customers")
    cust_list = [""] + cust_db["short_name"].tolist() if not cust_db.empty else [""]

    expected_cols = ['id', 'date_reported', 'date_resolved', 'customer_name', 'description', 'assignee', 'status', 'progress_pct', 'resolution_note', 'last_updated']
    if not df_issues.empty:
        for c in expected_cols:
            if c not in df_issues.columns: 
                df_issues[c] = None  
        
        # LOGIC TAB 7: Ép thẳng về định dạng Date của Pandas
        df_issues['date_reported'] = pd.to_datetime(df_issues['date_reported'], errors='coerce').dt.date
        df_issues['date_resolved'] = pd.to_datetime(df_issues['date_resolved'], errors='coerce').dt.date

    # 2. BỨC TRANH TOÀN CẢNH
    if not df_issues.empty:
        total_issues = len(df_issues)
        resolved_issues = len(df_issues[df_issues['status'].isin(['Resolved', 'Closed'])])
        open_issues = total_issues - resolved_issues

        i1, i2, i3 = st.columns(3)
        i1.markdown(f"<div class='card-3d bg-cost'><h3>SỰ CỐ ĐANG MỞ (OPEN)</h3><h1>{open_issues}</h1></div>", unsafe_allow_html=True)
        i2.markdown(f"<div class='card-3d bg-sales'><h3>ĐÃ GIẢI QUYẾT (RESOLVED)</h3><h1>{resolved_issues}</h1></div>", unsafe_allow_html=True)
        i3.markdown(f"<div class='card-3d bg-profit'><h3>TỔNG SỰ CỐ (TOTAL)</h3><h1>{total_issues}</h1></div>", unsafe_allow_html=True)
        st.divider()

    # 3. HEADER & XÓA ISSUE 
    c_i1, c_i2 = st.columns([7, 3])
    with c_i1:
        st.markdown("📋 **DANH SÁCH SỰ CỐ & TIẾN ĐỘ XỬ LÝ**")
        st.caption("💡 Mẹo 1: Kéo thả, copy-paste nhiều dòng từ Excel thoải mái. Mẹo 2: Nhấp đúp chuột vào ô để xem toàn bộ nội dung dài.")

    with c_i2:
        if not df_issues.empty:
            with st.popover("🗑️ XÓA ISSUE", use_container_width=True):
                st.markdown("**Chọn sự cố cần xóa**")
                issue_options = ["Chọn issue..."] + [f"[{r['id']}] - {r['customer_name']} - {str(r['description'])[:20]}..." for _, r in df_issues.iterrows()]
                selected_issue_del = st.selectbox("Danh sách:", issue_options, key="del_issue_select")
                
                if selected_issue_del != "Chọn issue...":
                    issue_id_to_del = int(selected_issue_del.split("]")[0].replace("[", ""))
                    st.warning(f"Đang chuẩn bị xóa Issue ID: **{issue_id_to_del}**")
                    del_pwd = st.text_input("Mật khẩu Admin", type="password", key="pwd_del_issue")
                    
                    if st.button("🔥 XÁC NHẬN XÓA", type="primary", use_container_width=True):
                        if del_pwd == "admin": 
                            try:
                                supabase.table("crm_issues").delete().eq("id", issue_id_to_del).execute()
                                st.cache_data.clear()
                                st.success(f"✅ Đã xóa thành công!")
                                time.sleep(1)
                                st.rerun()
                            except Exception as e: st.error(f"Lỗi xóa: {e}")
                        else: st.error("Sai mật khẩu!")

    # 4. CHIA TAB VÀ HIỂN THỊ
    # Nếu bảng trống, khởi tạo 1 DataFrame rỗng để user có thể nhập mới luôn
    if df_issues.empty:
        df_issues = pd.DataFrame(columns=expected_cols)

    tab_open, tab_resolved = st.tabs(["🔴 SỰ CỐ ĐANG MỞ (OPEN / IN PROGRESS)", "🟢 ĐÃ GIẢI QUYẾT (RESOLVED / CLOSED)"])
    
    def render_issue_table(df_subset, tab_key):
        # Thiết lập lại Index để đồng bộ số đếm chuẩn xác
        df_subset = df_subset.reset_index(drop=True)
        df_edit_issue = df_subset[expected_cols].copy()
        
        # ẨN HOÀN TOÀN CỘT ID KHỎI BẢNG GIAO DIỆN
        if 'id' in df_edit_issue.columns:
            df_edit_issue = df_edit_issue.drop(columns=['id'])
            
        # TẠO CỘT SỐ THỨ TỰ (NO) VÀO VỊ TRÍ ĐẦU TIÊN
        df_edit_issue.insert(0, "No", range(1, len(df_edit_issue) + 1))
        
        edited_issues = st.data_editor(
            df_edit_issue, 
            use_container_width=True, 
            hide_index=True, 
            height=600, # TĂNG CHIỀU CAO ĐỂ KÉO CHUỘT THOẢI MÁI NHƯ EXCEL
            num_rows="dynamic", # KÍCH HOẠT NHẬP DỮ LIỆU NHƯ EXCEL
            column_config={
                "No": st.column_config.NumberColumn("No.", disabled=True, width=50), # Hiện cột Số thứ tự
                "date_reported": st.column_config.DateColumn("Ngày PS", width=90),
                "date_resolved": st.column_config.DateColumn("Ngày KT", width=90),
                "customer_name": st.column_config.SelectboxColumn("Khách hàng", options=cust_list, width=120),
                "description": st.column_config.TextColumn("Mô tả vấn đề", width="large"), 
                "assignee": st.column_config.TextColumn("Người phụ trách", width=110),
                "status": st.column_config.SelectboxColumn("Trạng thái", options=["Open", "In Progress", "Resolved", "Closed"], width=100),
                "progress_pct": st.column_config.SelectboxColumn(
                    "Tiến độ", 
                    options=["0% ⚪", "10% 🔴", "20% 🔴", "30% 🟠", "40% 🟠", "50% 🟡", "60% 🟡", "70% 🔵", "80% 🔵", "90% 🔵", "100% 🟢"], 
                    width=90
                ),
                "resolution_note": st.column_config.TextColumn("Tình hình / Ghi chú", width="large"),
                "last_updated": None # Ẩn cột này
            },
            key=f"editor_issues_{tab_key}"
        )

        st.markdown('<div style="text-align: right;">', unsafe_allow_html=True)
        if st.button("💾 LƯU TẤT CẢ (THÊM MỚI & CẬP NHẬT)", type="primary", key=f"btn_update_issues_{tab_key}"):
            with st.spinner("Đang hạch toán dữ liệu hàng loạt..."):
                try:
                    changes_made = False
                    inserts = []
                    updates = []
                    
                    for idx, row in edited_issues.iterrows():
                        def get_str(val): return str(val).strip() if pd.notna(val) else ""
                        
                        desc = get_str(row.get('description'))
                        if not desc or desc.lower() == 'nan': 
                            continue # Bỏ qua dòng trống
                            
                        dr_new = str(row['date_reported']) if pd.notna(row.get('date_reported')) else datetime.now().strftime('%Y-%m-%d')
                        dres_new = str(row['date_resolved']) if pd.notna(row.get('date_resolved')) else None
                        status_new = row.get('status') if pd.notna(row.get('status')) else "Open"
                        prog_new = row.get('progress_pct') if pd.notna(row.get('progress_pct')) else "0% ⚪"
                        
                        # Tự động Resolved nếu 100%
                        if "100%" in str(prog_new) and status_new not in ["Resolved", "Closed"]:
                            status_new = "Resolved"
                            if not dres_new:
                                dres_new = datetime.now().strftime('%Y-%m-%d')
                        
                        # THUẬT TOÁN ĐỊNH VỊ INDEX (Khắc phục triệt để lỗi KeyError: 2)
                        if idx in df_subset.index:
                            # ==========================================
                            # LOGIC CẬP NHẬT DÒNG CŨ
                            # ==========================================
                            orig_row = df_subset.loc[idx]
                            db_id = int(orig_row['id'])
                            old_dr = str(orig_row['date_resolved']) if pd.notna(orig_row['date_resolved']) else None

                            if (get_str(row.get('status')) != get_str(orig_row['status']) or
                                get_str(row.get('progress_pct')) != get_str(orig_row['progress_pct']) or
                                get_str(row.get('resolution_note')) != get_str(orig_row['resolution_note']) or
                                get_str(row.get('assignee')) != get_str(orig_row['assignee']) or
                                get_str(row.get('description')) != get_str(orig_row['description']) or
                                get_str(row.get('customer_name')) != get_str(orig_row['customer_name']) or
                                dres_new != old_dr):

                                payload_update = {
                                    "id_upd": db_id, # Lưu tạm ID để gọi lệnh Update
                                    "status": status_new, 
                                    "progress_pct": prog_new,
                                    "resolution_note": get_str(row.get('resolution_note')), 
                                    "assignee": get_str(row.get('assignee')),
                                    "description": desc,
                                    "customer_name": get_str(row.get('customer_name')),
                                    "date_resolved": dres_new,
                                    "date_reported": dr_new,
                                    "last_updated": datetime.now().isoformat()
                                }
                                updates.append(payload_update)
                        else:
                            # ==========================================
                            # LOGIC INSERT DÒNG MỚI (Dù là 1 hay 100 dòng)
                            # ==========================================
                            payload_insert = {
                                "date_reported": dr_new,
                                "date_resolved": dres_new,
                                "customer_name": row.get('customer_name', ''),
                                "description": desc,
                                "assignee": row.get('assignee', ''),
                                "status": status_new,
                                "progress_pct": prog_new,
                                "resolution_note": get_str(row.get('resolution_note')),
                                "last_updated": datetime.now().isoformat()
                            }
                            inserts.append(payload_insert)

                    # THỰC THI BATCH CHUNG ĐỂ TRÁNH QUÁ TẢI SERVER
                    if inserts:
                        chunk_size = 50
                        for k in range(0, len(inserts), chunk_size):
                            supabase.table("crm_issues").insert(inserts[k:k+chunk_size]).execute()
                        changes_made = True
                        
                        # ========================================================
                        # GỬI TELEGRAM CHO TỪNG SỰ CỐ MỚI PHÁT SINH
                        # ========================================================
                        for ins in inserts:
                            try:
                                msg = (
                                    f"🆕 <b>PHÁT SINH SỰ CỐ MỚI</b>\n\n"
                                    f"🏢 <b>Khách hàng:</b> {ins['customer_name']}\n"
                                    f"📝 <b>Vấn đề:</b> {ins['description']}\n"
                                    f"👤 <b>Phụ trách:</b> {ins['assignee']}\n"
                                    f"<i>👉 Yêu cầu PIC tiếp nhận và xử lý!</i>"
                                )
                                url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
                                requests.post(url, json={"chat_id": TELEGRAM_GROUP_ID, "text": msg, "parse_mode": "HTML"})
                            except: pass

                    if updates:
                        for u in updates:
                            upd_id = u.pop("id_upd")
                            supabase.table("crm_issues").update(u).eq("id", upd_id).execute()
                            
                            # ========================================================
                            # GỬI TELEGRAM CHO TỪNG SỰ CỐ ĐƯỢC CẬP NHẬT
                            # ========================================================
                            try:
                                msg = (
                                    f"🔔 <b>CẬP NHẬT TIẾN ĐỘ SỰ CỐ</b>\n\n"
                                    f"🏢 <b>Khách hàng:</b> {u['customer_name']}\n"
                                    f"📝 <b>Vấn đề:</b> {u['description']}\n"
                                    f"👤 <b>Phụ trách:</b> {u['assignee']}\n"
                                    f"📊 <b>Tiến độ mới:</b> {u['progress_pct']}\n"
                                    f"🏷 <b>Trạng thái:</b> {u['status']}\n"
                                    f"<i>👉 Vui lòng kiểm tra CRM để xem chi tiết!</i>"
                                )
                                url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
                                requests.post(url, json={"chat_id": TELEGRAM_GROUP_ID, "text": msg, "parse_mode": "HTML"})
                            except: pass
                        changes_made = True

                    if changes_made:
                        st.cache_data.clear()
                        st.success("✅ Đã Lưu tất cả thay đổi & Thêm mới thành công!")
                        time.sleep(1)
                        st.rerun()
                    else: st.info("Không có thay đổi nào được ghi nhận.")
                except Exception as e: st.error(f"Lỗi: {e}")
        st.markdown('</div>', unsafe_allow_html=True)

    with tab_open:
        if not df_issues.empty:
            df_open = df_issues[~df_issues['status'].isin(['Resolved', 'Closed'])]
        else:
            df_open = df_issues.copy() # Truyền bảng rỗng để nhập liệu
        render_issue_table(df_open, "open")

    with tab_resolved:
        if not df_issues.empty:
            df_resolved = df_issues[df_issues['status'].isin(['Resolved', 'Closed'])]
            render_issue_table(df_resolved, "resolved")
        else:
            st.info("Chưa có sự cố nào được giải quyết.")
# =============================================================================
# --- TAB 9: THEO DÕI ĐƠN HÀNG (INDEPENDENT PO TRACKING) ---
# =============================================================================
with t9:
    st.markdown("### 📋 HỆ THỐNG THEO DÕI ĐƠN HÀNG ĐỘC LẬP")
    
    # --- 1. LOAD DATA NỘI BỘ ---
    # Sử dụng bảng riêng 'crm_po_tracking' để tránh conflict
    df_po_track = load_data("crm_po_tracking", order_by="id", ascending=False)
    cust_db = load_data("crm_customers")
    cust_list = [""] + cust_db["short_name"].tolist() if not cust_db.empty else [""]

    # --- 2. GIAO DIỆN CÔNG CỤ ---
    c_tool1, c_tool2 = st.columns([7, 3])
    
    with c_tool2:
        with st.popover("📥 IMPORT DATA HÀNG LOẠT", use_container_width=True):
            st.markdown("**Tải lên file Excel/CSV**")
            up_po_csv = st.file_uploader("Chọn file", type=["xlsx", "csv"], key="up_po_track_bulk")
            if up_po_csv and st.button("🚀 XÁC NHẬN IMPORT"):
                try:
                    if up_po_csv.name.endswith('.csv'):
                        df_imp = pd.read_csv(up_po_csv).fillna("")
                    else:
                        df_imp = pd.read_excel(up_po_csv).fillna("")
                    
                    # Mapping chuẩn hóa dữ liệu
                    imp_data = df_imp.to_dict('records')
                    supabase.table("crm_po_tracking").insert(imp_data).execute()
                    st.success("✅ Đã import dữ liệu thành công!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Lỗi: {e}")

    # --- 3. BẢNG DỮ LIỆU CHÍNH (DATA EDITOR) ---
    st.info("💡 Bạn có thể copy-paste trực tiếp từ Excel vào bảng dưới đây.")
    
    # Định nghĩa cấu trúc cột theo yêu cầu
    expected_po_cols = [
        'customer', 'po_no', 'req_no', 'item_code', 'item_name', 
        'specs', 'qty', 'unit_price', 'total_price', 'po_docs', 'remark'
    ]

    # Chuẩn bị DataFrame hiển thị
    if df_po_track.empty:
        df_display_po = pd.DataFrame(columns=expected_po_cols)
    else:
        df_display_po = df_po_track.copy()
        # Loại bỏ các cột hệ thống để bảng sạch hơn
        if 'id' in df_display_po.columns: df_id_map = df_display_po['id'].tolist()
        df_display_po = df_display_po[expected_po_cols]

    # Thêm cột No
    df_display_po.insert(0, "No", range(1, len(df_display_po) + 1))

    edited_po_tab9 = st.data_editor(
        df_display_po,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        height=500,
        column_config={
            "No": st.column_config.NumberColumn("No", width=40, disabled=True),
            "customer": st.column_config.SelectboxColumn("Customer", options=cust_list, width=150),
            "qty": st.column_config.NumberColumn("Q'ty", format="%d"),
            "unit_price": st.column_config.NumberColumn("Unit Price", format="%d"),
            "total_price": st.column_config.NumberColumn("Total Price", format="%d"),
            "po_docs": st.column_config.LinkColumn("PO Docs", width=150, display_text="📂 Mở tài liệu"),
            "item_name": st.column_config.TextColumn("Item Name", width=200),
            "remark": st.column_config.TextColumn("Remark", width=200),
        },
        key="editor_po_tracking_tab9"
    )

    # --- 4. XỬ LÝ LƯU & UPLOAD & TELEGRAM ---
    st.markdown("---")
    col_save1, col_save2 = st.columns([3, 7])
    
    with col_save1:
        # Chức năng upload file cho dòng được chọn
        st.markdown("**📂 Upload tài liệu cho PO**")
        target_po_idx = st.number_input("Nhập 'No' của dòng muốn upload", min_value=1, max_value=len(edited_po_tab9) if not edited_po_tab9.empty else 1, step=1)
        up_files_po = st.file_uploader("Đính kèm (Ảnh, PDF, Excel, Word...)", accept_multiple_files=True, key="up_po_docs_t9")

    with col_save2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("💾 LƯU THAY ĐỔI & GỬI THÔNG BÁO TELEGRAM", type="primary", use_container_width=True):
            try:
                # 1. Xử lý upload Drive nếu có file
                doc_link = ""
                if up_files_po:
                    with st.spinner("Đang tải tài liệu lên Drive..."):
                        po_ref = edited_po_tab9.iloc[target_po_idx-1]['po_no'] or "UNNAMED_PO"
                        path_list = ["PO_TRACKING_DOCS", str(po_ref)]
                        srv = get_drive_service()
                        folder_id = get_or_create_folder_hierarchy(srv, path_list, ROOT_FOLDER_ID)
                        doc_link = f"https://drive.google.com/drive/folders/{folder_id}"
                        for f in up_files_po:
                            upload_to_drive_structured(f, path_list, f.name)

                # 2. Xử lý lưu Database
                # Xóa dữ liệu cũ của bảng này để ghi đè (đảm bảo tính độc lập)
                supabase.table("crm_po_tracking").delete().neq("id", 0).execute()
                
                new_records = []
                for i, row in edited_po_tab9.iterrows():
                    # Gán link docs cho dòng được chọn
                    if i == (target_po_idx - 1) and doc_link:
                        row['po_docs'] = doc_link
                    
                    # Chỉ lưu dòng có thông tin
                    if row['customer'] and row['po_no']:
                        data_row = {k: row[k] for k in expected_po_cols}
                        new_records.append(data_row)
                
                if new_records:
                    supabase.table("crm_po_tracking").insert(new_records).execute()
                    
                    # 3. Gửi thông báo Telegram (Sử dụng thuật toán Tab 7/8)
                    last_row = new_records[-1] # Thông báo cho đơn hàng mới nhất/vừa cập nhật
                    msg = (
                        f"📦 <b>CẬP NHẬT ĐƠN HÀNG (PO)</b>\n\n"
                        f"👤 <b>Khách hàng:</b> {last_row['customer']}\n"
                        f"📄 <b>Số PO:</b> {last_row['po_no']}\n"
                        f"💰 <b>Tổng tiền:</b> {fmt_num(last_row['total_price'])} VND\n"
                        f"📅 <b>Ngày nhận:</b> {datetime.now().strftime('%d/%m/%Y')}\n"
                        f"📝 <b>Ghi chú:</b> {last_row['remark']}\n\n"
                        f"<i>👉 Xem chi tiết tại Tab 9 hệ thống CRM!</i>"
                    )
                    url_tele = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
                    requests.post(url_tele, json={"chat_id": TELEGRAM_GROUP_ID, "text": msg, "parse_mode": "HTML"})
                    
                    st.success("✅ Đã lưu dữ liệu và gửi thông báo Telegram!")
                    time.sleep(1)
                    st.rerun()
            except Exception as e:
                st.error(f"Lỗi: {e}")

# =============================================================================
# --- KẾT THÚC TAB 9 ---
